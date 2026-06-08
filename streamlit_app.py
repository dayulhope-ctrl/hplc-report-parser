# -*- coding: utf-8 -*-
import sys, json, os, tempfile, math, io, shutil
from pathlib import Path
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR   = Path(__file__).parent
CONFIG_DIR = BASE_DIR / "config"
AA_DIR     = Path(r"C:\Users\user\Downloads\파워라센 아미노산 함량")

sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(AA_DIR))

st.set_page_config(page_title="면적값 자동파싱 시스템", layout="wide", initial_sidebar_state="collapsed")

# ══════════════════════════════════════════════════════════════════
# 분석 함수들
# ══════════════════════════════════════════════════════════════════

def _run_id(product, spa_file, spb_file, spc_file):
    from core.id_parser import parse_id_pdf, compute_rrt, judge_rrt
    from core.excel_writer import write_id_result

    form_type   = "환" if product == "환제" else "현"
    config_file = CONFIG_DIR / ("csw_pill.json" if form_type == "환" else "csw_hyeon.json")
    with open(config_file, encoding="utf-8") as f:
        config = json.load(f)

    compounds = config["compounds_id"]
    ref_comp  = config["id_reference_compound"]
    tolerance = config["id_rrt_tolerance"]

    all_sample_data = {}
    with st.spinner("PDF 파싱 중..."):
        for label, upfile in [("A", spa_file), ("B", spb_file), ("C", spc_file)]:
            if upfile is None:
                continue
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp.write(upfile.read()); tmp_path = tmp.name
            try:
                parsed = parse_id_pdf(tmp_path)
                for sn, sdata in parsed.items():
                    all_sample_data[sn] = compute_rrt(sdata, ref_comp)
            finally:
                os.unlink(tmp_path)

    if not all_sample_data:
        st.error("PDF에서 데이터를 추출하지 못했습니다."); return

    sample_names = list(all_sample_data.keys())
    st.success(f"추출 완료: {len(sample_names)}개 시료 — {', '.join(sample_names)}")

    result_rows = []
    for comp_cfg in compounds:
        comp_name = comp_cfg["name"]
        rrt_exp   = comp_cfg["rrt_expected"]
        is_ref    = comp_cfg.get("is_reference", False)
        sample_results = {}
        for sn in sample_names:
            sdata   = all_sample_data[sn]
            matched = next((sdata[k] for k in sdata if k.startswith(comp_name)), None)
            if matched:
                rrt_c  = matched.get("rrt_calc")
                result = "기준" if is_ref else (judge_rrt(rrt_c, rrt_exp, tolerance) if rrt_c else "N/A")
                sample_results[sn] = {"rt": matched.get("rt"), "sn": matched.get("sn"),
                                      "rrt_calc": rrt_c, "result": result}
            else:
                sample_results[sn] = {"rt": None, "sn": None, "rrt_calc": None, "result": "미검출"}
        result_rows.append({"compound": comp_name, "transitions": comp_cfg["transitions"],
                            "rrt_expected": rrt_exp, "samples": sample_results})

    summary = []
    for row in result_rows:
        r = {"성분": row["compound"], "RRT 기준": row["rrt_expected"]}
        for sn in sample_names: r[sn] = row["samples"][sn].get("result", "-")
        summary.append(r)

    df = pd.DataFrame(summary)
    def color_result(val):
        if val in ("적합","기준"): return "background-color:#C6EFCE"
        if val == "부적합":        return "background-color:#FFC7CE"
        return ""
    st.dataframe(df.style.applymap(color_result, subset=sample_names),
                 use_container_width=True, hide_index=True)

    fails = [f"{r['성분']}({sn})" for r in summary for sn in sample_names if r.get(sn) == "부적합"]
    if fails: st.warning(f"부적합: {', '.join(fails)}")
    else:     st.success("전 항목 적합")

    excel = write_id_result(result_rows, lot_info=product)
    st.download_button("📥 결과 엑셀 다운로드", data=excel,
                       file_name=f"{product}_확인결과.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       type="primary", use_container_width=True)


def _run_id_csv(product, sst_file, spa_file, spb_file, spc_file):
    from core.id_csv_parser import parse_sp_csv, parse_sst_csv
    from core.id_csv_excel_writer import write_id_csv_result

    with st.spinner("CSV 파싱 중..."):
        sst_data = parse_sst_csv(sst_file) if sst_file else {}
        sp_a     = parse_sp_csv(spa_file)  if spa_file else {}
        sp_b     = parse_sp_csv(spb_file)  if spb_file else {}
        sp_c     = parse_sp_csv(spc_file)  if spc_file else {}

    lot = sp_a.get("lot_name") or sp_b.get("lot_name") or sp_c.get("lot_name") or "결과"
    st.success(f"파싱 완료 — Lot: {lot}")

    from core.id_csv_excel_writer import HYEON_SP_ORDER
    order = HYEON_SP_ORDER if product == "현탁제" else None
    excel_bytes = write_id_csv_result(sst_data, sp_a, sp_b, sp_c, compound_order=order)
    st.download_button("📥 결과 엑셀 다운로드", data=excel_bytes,
                       file_name=f"{product}_확인결과_{lot}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       type="primary", use_container_width=True)


def _run_as(product, as_files):
    from core.csv_utils import merge_parsed
    if product == "현탁제":
        from core.as_hyeon_csv_parser import group_hyeon_sp_by_lot as _group, get_hyeon_std_stats as _stats
        from core.as_hyeon_excel_writer import write_hyeon_result as _write
    else:
        from core.as_csv_parser import group_sp_by_lot as _group, get_std_stats as _stats
        from core.as_excel_writer import write_as_result as _write

    with st.spinner("CSV 파싱 중..."):
        parsed     = merge_parsed(as_files, prefer="area")
        lot_groups = _group(parsed)

    if not parsed:
        st.error("CSV에서 데이터를 추출하지 못했습니다."); return

    lot_names = list(lot_groups.keys())
    st.success(f"추출 완료 — STD 6런 / SP Lot: {', '.join(lot_names)}")

    # 미리보기
    preview = []
    for comp, cdata in parsed.items():
        stats = _stats(cdata)
        preview.append({
            "성분": comp,
            "STD 런": stats.get("count", 0),
            "STD 평균 Area": stats.get("avg_area"),
            "%RSD": stats.get("rsd"),
        })
    st.dataframe(pd.DataFrame(preview), use_container_width=True, hide_index=True)

    excel_bytes = _write(parsed, lot_groups)
    st.download_button("📥 결과 엑셀 다운로드", data=excel_bytes,
                       file_name=f"{product}_함량결과.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       type="primary", use_container_width=True)


def _run_amino_acid(std_file, sp_file):
    TEMPLATE = AA_DIR / "아미노산_함량_분석_양식.xlsx"
    if not TEMPLATE.exists():
        st.error(f"양식 파일 없음: {TEMPLATE}"); return

    from parse_pdf import parse_both, validate_data
    from fill_form import fill_form, find_std_area_cells as _find_std
    from template_creator import build_lot_sheet
    import json as _json

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        std_path = tmpdir/"STD.pdf"; std_path.write_bytes(std_file.read())
        sp_path  = tmpdir/"SP.pdf";  sp_path.write_bytes(sp_file.read())

        with st.spinner("PDF 파싱 중..."):
            try:    data = parse_both(str(std_path), str(sp_path))
            except Exception as e: st.error(f"파싱 오류: {e}"); return

        warns = validate_data(data)
        if warns:
            with st.expander(f"경고 {len(warns)}건"):
                [st.warning(w) for w in warns[:30]]
        else:
            st.success("검증 통과")

        lot_names = sorted(data["lots"].keys())
        st.info(f"감지된 Lot: {', '.join(lot_names)}")

        with st.spinner("엑셀 생성 중..."):
            out_path = tmpdir/"결과.xlsx"
            shutil.copy2(TEMPLATE, out_path)
            wb = openpyxl.load_workbook(str(out_path))

            std_avg_cells = {}
            if "표준액(면적)" in wb.sheetnames:
                for aa, addrs in _find_std(wb["표준액(면적)"]).items():
                    if addrs:
                        last = addrs[-1]
                        col = ''.join(filter(str.isalpha, last))
                        row = int(''.join(filter(str.isdigit, last))) + 1
                        std_avg_cells[aa] = f"{col}{row}"

            for lot in lot_names:
                if lot not in wb.sheetnames:
                    build_lot_sheet(wb, lot, std_avg_cells)
            wb.save(str(out_path))

            json_tmp = tmpdir/"_data.json"
            with open(json_tmp,"w",encoding="utf-8") as f:
                _json.dump(data, f, ensure_ascii=False)

            try:    fill_form(str(out_path), str(json_tmp), str(out_path))
            except Exception as e: st.error(f"엑셀 입력 오류: {e}"); return

        st.success(f"완료! {len(lot_names)}개 Lot 생성")
        excel_bytes = out_path.read_bytes()

    st.download_button("📥 결과 엑셀 다운로드", data=excel_bytes,
                       file_name="아미노산_함량_결과.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       type="primary", use_container_width=True)


# ══════════════════════════════════════════════════════════════════
# 세션 상태 초기화
# ══════════════════════════════════════════════════════════════════
if "product"   not in st.session_state: st.session_state.product   = None
if "test_type" not in st.session_state: st.session_state.test_type = None

# ══════════════════════════════════════════════════════════════════
# UI — 헤더
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
div.stButton > button { height:55px; font-size:16px; font-weight:bold; border-radius:10px; }
</style>
""", unsafe_allow_html=True)

st.title("🔬 면적값 자동파싱 시스템")
st.caption("PDF를 업로드하면 자동으로 데이터를 추출하여 엑셀로 제공합니다")
st.divider()

# ══════════════════════════════════════════════════════════════════
# STEP 1 — 제품 선택
# ══════════════════════════════════════════════════════════════════
st.subheader("① 제품 선택")
col1, col2, col3 = st.columns(3)

with col1:
    if st.button("💊  환제", use_container_width=True,
                 type="primary" if st.session_state.product=="환제" else "secondary"):
        st.session_state.product="환제"; st.session_state.test_type=None; st.rerun()
with col2:
    if st.button("🧪  현탁제", use_container_width=True,
                 type="primary" if st.session_state.product=="현탁제" else "secondary"):
        st.session_state.product="현탁제"; st.session_state.test_type=None; st.rerun()
with col3:
    if st.button("🔬  파워라센 아미노산", use_container_width=True,
                 type="primary" if st.session_state.product=="아미노산" else "secondary"):
        st.session_state.product="아미노산"; st.session_state.test_type=None; st.rerun()

if not st.session_state.product:
    st.stop()

st.divider()

# ══════════════════════════════════════════════════════════════════
# STEP 2 — 시험 항목 선택 (아미노산 제외)
# ══════════════════════════════════════════════════════════════════
product = st.session_state.product

if product in ("환제", "현탁제"):
    st.subheader(f"② 시험 항목 선택  〔우황청심원 {product}〕")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📋  확인 (ID)", use_container_width=True,
                     type="primary" if st.session_state.test_type=="확인" else "secondary"):
            st.session_state.test_type="확인"; st.rerun()
    with col2:
        if st.button("⚗️  함량 (AS)", use_container_width=True,
                     type="primary" if st.session_state.test_type=="함량" else "secondary"):
            st.session_state.test_type="함량"; st.rerun()

    if not st.session_state.test_type:
        st.stop()
    st.divider()

test_type = st.session_state.test_type

# ══════════════════════════════════════════════════════════════════
# STEP 3 — PDF 업로드
# ══════════════════════════════════════════════════════════════════

if product == "아미노산":
    st.subheader("② PDF 파일 업로드  〔아미노산 함량〕")
    st.caption("STD / SP 파일을 한 번에 선택하세요 (파일명에 STD, SP 포함)")
    uploaded_aa = st.file_uploader("PDF 파일 선택 (복수 선택 가능)",
                                   type="pdf", accept_multiple_files=True, key="aa_multi")
    std_file = sp_file = None
    others = []
    for f in (uploaded_aa or []):
        n = f.name.upper()
        if "STD" in n: std_file = f
        elif "SP" in n: sp_file = f
        else: others.append(f)
    # STD가 아닌 미분류 파일 → SP로 처리
    if sp_file is None and others:
        sp_file = others[0]
    if uploaded_aa:
        cols = st.columns(2)
        cols[0].success(f"✓ {std_file.name}" if std_file else "—"); cols[0].caption("STD")
        cols[1].success(f"✓ {sp_file.name}"  if sp_file  else "—"); cols[1].caption("SP")
    if st.button("▶  분석 실행", type="primary",
                 disabled=not(std_file and sp_file), use_container_width=True):
        _run_amino_acid(std_file, sp_file)

elif test_type == "확인":
    if product == "환제":
        st.subheader(f"③ CSV 파일 업로드  〔우황청심원 {product} 확인〕")
        st.caption("SST / SP_A / SP_B / SP_C 파일을 한 번에 선택하세요 (파일명에 SST, SP_A, SP_B, SP_C 포함)")
        uploaded = st.file_uploader("CSV 파일 선택 (복수 선택 가능)",
                                    type="csv", accept_multiple_files=True,
                                    key="id_csv_multi")
        # 파일명으로 자동 분류
        sst_file = spa_file = spb_file = spc_file = None
        for f in (uploaded or []):
            n = f.name.upper()
            if "SST"  in n: sst_file = f
            elif "SP_A" in n or "SPA" in n: spa_file = f
            elif "SP_B" in n or "SPB" in n: spb_file = f
            elif "SP_C" in n or "SPC" in n: spc_file = f
        # 업로드 현황 표시
        if uploaded:
            cols = st.columns(4)
            for col, label, f in zip(cols,
                ["SST", "SP_A", "SP_B", "SP_C"],
                [sst_file, spa_file, spb_file, spc_file]):
                col.success(f"✓ {f.name}" if f else "—")
                col.caption(label)
        if st.button("▶  분석 실행", type="primary",
                     disabled=not any([spa_file, spb_file, spc_file]),
                     use_container_width=True):
            _run_id_csv(product, sst_file, spa_file, spb_file, spc_file)
    else:
        st.subheader(f"③ CSV 파일 업로드  〔우황청심원 {product} 확인〕")
        st.caption("SST / SP_A / SP_B / SP_C 파일을 한 번에 선택하세요 (파일명에 SST, SP_A, SP_B, SP_C 포함)")
        uploaded = st.file_uploader("CSV 파일 선택 (복수 선택 가능)",
                                    type="csv", accept_multiple_files=True,
                                    key="id_csv_multi2")
        sst_file = spa_file = spb_file = spc_file = None
        for f in (uploaded or []):
            n = f.name.upper()
            if "SST"  in n: sst_file = f
            elif "SP_A" in n or "SPA" in n: spa_file = f
            elif "SP_B" in n or "SPB" in n: spb_file = f
            elif "SP_C" in n or "SPC" in n: spc_file = f
        if uploaded:
            cols = st.columns(4)
            for col, label, f in zip(cols,
                ["SST", "SP_A", "SP_B", "SP_C"],
                [sst_file, spa_file, spb_file, spc_file]):
                col.success(f"✓ {f.name}" if f else "—")
                col.caption(label)
        if st.button("▶  분석 실행", type="primary",
                     disabled=not any([spa_file, spb_file, spc_file]),
                     use_container_width=True):
            _run_id_csv(product, sst_file, spa_file, spb_file, spc_file)

elif test_type == "함량":
    st.subheader(f"③ CSV 파일 업로드  〔우황청심원 {product} 함량〕")
    st.caption("통합 CSV 1개 또는 분리된 CSV 여러 개를 한 번에 선택하세요")
    as_files = st.file_uploader("AS 데이터 CSV (복수 선택 가능)",
                                type=["csv"], accept_multiple_files=True, key="as_csv")
    if as_files:
        for f in as_files:
            st.caption(f"✓ {f.name}")
    if st.button("▶  분석 실행", type="primary",
                 disabled=not as_files, use_container_width=True):
        _run_as(product, as_files)
