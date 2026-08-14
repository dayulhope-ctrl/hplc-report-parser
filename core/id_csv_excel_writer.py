# -*- coding: utf-8 -*-
"""
환제 확인시험 엑셀 라이터
원본 AS.csv 양식 그대로 재현:
  - 각 화합물 그룹 (3 transition) × (RT, S/N) = 6 데이터 열
  - 구분 열 없이 데이터 직접 출력
  - SST 섹션은 구분 열(샘플명) 포함
  - 물질명 기반 정확한 열 매핑
"""

import io, re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 색상 ──────────────────────────────────────────────────────────
YELLOW   = PatternFill("solid", fgColor="FFFFE1")
BLUE_H   = PatternFill("solid", fgColor="BDD7EE")
GRAY     = PatternFill("solid", fgColor="D9D9D9")
WHITE    = PatternFill("solid", fgColor="FFFFFF")
TITLE_F  = PatternFill("solid", fgColor="2E75B6")
GREEN    = PatternFill("solid", fgColor="C6EFCE")
RED_F    = PatternFill("solid", fgColor="FFC7CE")

FS_SST = 14   # SST 섹션 글꼴 크기
FS_SP  = 11   # Value(SP) 섹션 글꼴 크기

AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left",   vertical="center")
THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 데이터 열 시작
SST_LABEL_COL = 1   # SST 구분 열 (A)
SST_DATA_COL  = 2   # SST 데이터 시작 열 (B)
SP_DATA_COL   = 2   # SP 데이터 시작 열 (B) — col A 비움
N_TRANS       = 3


def _c(ws, row, col, value="", fill=None, bold=False,
       num_fmt=None, align="center", font_color=None, fs=FS_SP):
    from openpyxl.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return cell
    cell.value = value
    cell.font  = Font(name="맑은 고딕", size=fs, bold=bold,
                      color=font_color if font_color else "000000")
    cell.alignment = AC if align == "center" else AL
    cell.border    = BORDER
    if fill:    cell.fill = fill
    if num_fmt: cell.number_format = num_fmt
    return cell


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _title_bar(ws, row, text, c1, c2, fs=FS_SP):
    _merge(ws, row, c1, row, c2)
    cell = ws.cell(row=row, column=c1)
    cell.value     = text
    cell.font      = Font(name="맑은 고딕", size=fs, bold=True, color="FFFFFF")
    cell.fill      = TITLE_F
    cell.alignment = AL
    cell.border    = BORDER
    ws.row_dimensions[row].height = 18


def _base_name(name):
    """"Platycodin-D_1223.7" → "Platycodin-D" """
    return re.sub(r'_[\d.]+$', '', name)


# ── 양식 고정 transition 순서 ─────────────────────────────────────
TRANSITION_ORDER = {
    # 환제 SP_A
    "Platycodin-D":             ["Platycodin-D_1223.7",            "Platycodin-D_681.1",            "Platycodin-D_469.2"],
    "Ginsenoside Rb1":          ["Ginsenoside Rb1_945.3",          "Ginsenoside Rb1_221.2",         "Ginsenoside Rb1_178.8"],
    "BoRy":                     ["BoRy_311.3",                     "BoRy_293.2",                    "BoRy_55.1"],
    "Bilirubin":                ["Bilirubin_285.2",                "Bilirubin_241.2",               "Bilirubin_213.0"],
    # 환제 SP_B
    "Prim-O-glucosylcimifugin": ["Prim-O-glucosylcimifugin_307.2", "Prim-O-glucosylcimifugin_259.0","Prim-O-glucosylcimifugin_235.1"],
    "PH":                       ["PH_314.1",                       "PH_298.8",                      "PH_271.0"],
    "SY":                       ["SY_147.1",                       "SY_121.1",                      "SY_77.1"],
    "Saikosaponin A":           ["Saikosaponin A_617.6",           "Saikosaponin A_145.0",          "Saikosaponin A_101.0"],
    "6-Gingerol":               ["6-Gingerol_193.0",               "6-Gingerol_99.0",               "6-Gingerol_57.0"],
    "Atractylenolide III":      ["Atractylenolide III_203.1",      "Atractylenolide III_187.1",     "Atractylenolide III_83.1"],
    "Ligustilide":              ["Ligustilide_115.1",              "Ligustilide_77.0",              "Ligustilide_51.1"],
    # 환제 SP_C
    "Amygdalin":                ["Amygdalin_323.0",                "Amygdalin_221.0",               "Amygdalin_161.0"],
    "Paeoniflorin":             ["Paeoniflorin_327.2",             "Paeoniflorin_164.9",            "Paeoniflorin_121.1"],
    "Baicalin":                 ["Baicalin_123.0",                 "Baicalin_103.0",                "Baicalin_94.9"],
    "Glycyrrhizin":             ["Glycyrrhizin_350.7",             "Glycyrrhizin_193.0",            "Glycyrrhizin_112.9"],
    "Decursin":                 ["Decursin_247.2",                 "Decursin_128.0",                "Decursin_115.0"],
    # 현탁액 추가
    "Ginsenoside_Rg1":          ["Ginsenoside_Rg1_799.6",         "Ginsenoside_Rg1_637.4",         "Ginsenoside_Rg1_59.0"],
    "Saikosaponin-A":           ["Saikosaponin-A_617.6",          "Saikosaponin-A_145.0",          "Saikosaponin-A_101.0"],
}


def _group_transitions(transitions):
    """
    CSV 순서 무관하게 TRANSITION_ORDER 고정 순서로 그룹핑.
    [(base_name, [t1, t2, t3]), ...]
    """
    # transition 이름 → 데이터 매핑
    trans_map = {t["name"]: t for t in transitions}

    groups, seen = [], set()
    for t in transitions:
        base = _base_name(t["name"])
        if base in seen:
            continue
        seen.add(base)

        if base in TRANSITION_ORDER:
            ordered = []
            for name in TRANSITION_ORDER[base]:
                if name in trans_map:
                    ordered.append(trans_map[name])
                else:
                    ordered.append({"name": name, "rt": None, "sn": None})
        else:
            # 미등록 화합물은 CSV 순서 그대로
            ordered = [t2 for t2 in transitions if _base_name(t2["name"]) == base]

        groups.append((base, ordered))
    return groups


def _set_col_widths(ws, show_rrt=False):
    ws.column_dimensions["A"].width = 20
    for col in range(2, 9):    # B~H
        ws.column_dimensions[get_column_letter(col)].width = 14
    if show_rrt:
        for col in range(9, 12):   # I~K: RRT / 기준RRT / 판정
            ws.column_dimensions[get_column_letter(col)].width = 10


# ════════════════════════════════════════════════════════════════
# 메인
# ════════════════════════════════════════════════════════════════
# 현탁액 전용 화합물 출력 순서
HYEON_SP_ORDER = [
    "Amygdalin", "Ginsenoside_Rg1", "Prim-O-glucosylcimifugin",
    "Baicalin", "Glycyrrhizin", "Decursin",
    "Paeoniflorin", "Saikosaponin-A", "Atractylenolide III",
    "PH", "Platycodin-D", "6-Gingerol", "Ligustilide", "Bilirubin",
]


# ── RRT 기준값 (±10%) ─────────────────────────────────────────────
RRT_REF = {
    "환제": {
        "Amygdalin": 0.3,
        "Paeoniflorin": 0.4,
        "Prim-O-glucosylcimifugin": 0.4,
        "PH": 0.5,
        "Baicalin": 0.7,
        "SY": 0.7,
        "Platycodin-D": 0.8,
        "Ginsenoside Rb1": 1.0,
        "Glycyrrhizin": 1.1,
        "Saikosaponin A": 1.2,
        "6-Gingerol": 1.2,
        "Atractylenolide III": 1.3,
        "Ligustilide": 1.4,
        "Decursin": 1.6,
        "BoRy": 1.7,
        "Bilirubin": 1.8,
    },
    "현탁제": {
        "Amygdalin": 0.4,
        "Paeoniflorin": 0.5,
        "Prim-O-glucosylcimifugin": 0.6,
        "PH": 0.7,
        "Baicalin": 0.9,
        "Ginsenoside_Rg1": 1.0,
        "Platycodin-D": 1.2,
        "Glycyrrhizin": 1.6,
        "6-Gingerol": 1.7,
        "Saikosaponin-A": 1.7,
        "Atractylenolide III": 1.8,
        "Ligustilide": 2.0,
        "Decursin": 2.2,
        "Bilirubin": 2.5,
    },
}
REFERENCE_COMPOUND = {
    "환제": "Ginsenoside Rb1",
    "현탁제": "Ginsenoside_Rg1",
}

def _build_sheet(wb, ws_title, sst_data, sp_files, lot_idx, compound_order, form_type=None):
    """lot_idx번째 샘플 데이터로 시트 1장 생성."""
    ws = wb.create_sheet(title=ws_title[:31])
    show_rrt = bool(form_type)
    _set_col_widths(ws, show_rrt=show_rrt)

    # 기준화합물 RT 추출
    ref_rt = None
    if form_type:
        ref_comp = REFERENCE_COMPOUND.get(form_type)
        if ref_comp:
            for _sp in sp_files:
                if not _sp: continue
                _tlist = _sp.get("all_transitions") or [_sp.get("transitions", [])]
                _sample = _tlist[lot_idx] if lot_idx < len(_tlist) else _tlist[0]
                for _t in _sample:
                    if _base_name(_t["name"]) == ref_comp:
                        ref_rt = _t.get("rt")
                        break
                if ref_rt: break

    row = 1
    cs       = SP_DATA_COL
    last_col = cs + N_TRANS * 2 - 1 + (3 if show_rrt else 0)

    if sst_data:
        row = _write_sst_section(ws, row, sst_data)
        row += 1

    _title_bar(ws, row, "● Value", cs, cs + 5)
    row += 1
    sp_section_start = row

    if compound_order:
        all_groups = {}
        for sp_data in sp_files:
            if not sp_data:
                continue
            all_trans_list = sp_data.get("all_transitions") or [sp_data.get("transitions", [])]
            single = [all_trans_list[lot_idx]] if lot_idx < len(all_trans_list) else [all_trans_list[0]]
            for base_name, trans_list in _group_transitions(all_trans_list[0]):
                all_groups[base_name] = (trans_list, single)

        for base_name in compound_order:
            if base_name not in all_groups:
                continue
            trans_list, single = all_groups[base_name]
            _ref_rrt = RRT_REF.get(form_type, {}).get(base_name) if form_type else None
            row = _write_compound_block(ws, row, trans_list, single,
                                        ref_rts=[ref_rt], ref_rrt=_ref_rrt)
    else:
        for sp_data in sp_files:
            if not sp_data:
                continue
            all_trans_list = sp_data.get("all_transitions") or [sp_data.get("transitions", [])]
            single_sp = dict(sp_data)
            single_sp["all_transitions"] = [all_trans_list[lot_idx]] if lot_idx < len(all_trans_list) else [all_trans_list[0]]
            row = _write_sp_compounds(ws, row, single_sp, form_type=form_type, ref_rt=ref_rt)

    if row > sp_section_start:
        _apply_outer_border(ws, sp_section_start, cs, row - 1, last_col)


def write_id_csv_result(sst_data: dict,
                        sp_a: dict, sp_b: dict, sp_c: dict,
                        compound_order: list = None,
                        form_type: str = None) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sp_files = [s for s in [sp_a, sp_b, sp_c] if s]

    # 모든 SP 파일에서 샘플 이름 목록 수집 (가장 많은 파일 기준)
    all_sample_names = []
    for sp in sp_files:
        names = sp.get("sample_names") or [sp.get("sample_name", "결과")]
        if len(names) > len(all_sample_names):
            all_sample_names = names

    if not all_sample_names:
        all_sample_names = ["결과"]

    for lot_idx, sample_name in enumerate(all_sample_names):
        # 시트 이름: 파일명에서 확장자 및 -A/-B/-C 제거
        import re as _re
        sheet_title = _re.sub(r'\.d$', '', sample_name, flags=_re.IGNORECASE)
        sheet_title = _re.sub(r'-[A-Ca-c]$', '', sheet_title)
        _build_sheet(wb, sheet_title, sst_data, sp_files, lot_idx, compound_order, form_type=form_type)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── SST 섹션 (RT만, 번호행 1·2·3 + 통계 + 판정) ─────────────────
def _write_sst_section(ws, row, sst_data):
    comp_names = sst_data.get("compound_names", [])
    n          = len(comp_names)
    last_col   = SST_LABEL_COL + n   # 구분(1) + 화합물 수

    _title_bar(ws, row, "● 시스템적합성 (SST)", SST_LABEL_COL, last_col, fs=FS_SST)
    row += 1

    # 헤더: 구분 | Comp1 | Comp2 | Comp3
    _c(ws, row, SST_LABEL_COL, "구분", fill=BLUE_H, bold=True, fs=FS_SST)
    for i, name in enumerate(comp_names):
        _c(ws, row, SST_DATA_COL + i, name, fill=BLUE_H, bold=True, fs=FS_SST)
    row += 1

    # SST 데이터 행 (번호 1·2·3) — RT만
    sst_rows   = sst_data.get("sst", [])
    data_start = row
    for idx, row_data in enumerate(sst_rows, 1):
        _c(ws, row, SST_LABEL_COL, idx, fill=WHITE, fs=FS_SST)
        for i, name in enumerate(comp_names):
            rt = row_data.get(name, {}).get("rt")
            _c(ws, row, SST_DATA_COL + i, (round(rt, 3) if rt is not None else None), fill=YELLOW, num_fmt="0.000", fs=FS_SST)
        row += 1
    data_end = row - 1

    # 통계 행
    for stat_label, stat_func in [("평균", "AVERAGE"),
                                   ("표준편차", "STDEV"),
                                   ("유지시간 RSD %", None)]:
        _c(ws, row, SST_LABEL_COL, stat_label, fill=GRAY, bold=True, fs=FS_SST)
        for i in range(n):
            col_letter = get_column_letter(SST_DATA_COL + i)
            rng = f"{col_letter}{data_start}:{col_letter}{data_end}"
            if stat_func:
                formula = f"={stat_func}({rng})"
                fmt = "0.000"
            else:
                avg_row = row - 2
                std_row = row - 1
                cl = get_column_letter(SST_DATA_COL + i)
                formula = f"={cl}{std_row}/{cl}{avg_row}*100"
                fmt = "0.0"
            cell = ws.cell(row, SST_DATA_COL + i, formula)
            cell.fill = GRAY; cell.border = BORDER
            cell.font = Font(name="맑은 고딕", size=FS_SST, bold=True)
            cell.alignment = AC; cell.number_format = fmt
        row += 1

    # 판정 행 (RSD ≤ 2.0% → 적합)
    rsd_row = row - 1
    _c(ws, row, SST_LABEL_COL, "판정 (2.0 % 이하)", fill=GRAY, bold=True, fs=FS_SST)
    for i in range(n):
        cl = get_column_letter(SST_DATA_COL + i)
        cell = ws.cell(row, SST_DATA_COL + i,
                       f'=IF({cl}{rsd_row}<=2.0,"적합","부적합")')
        cell.fill = GRAY; cell.border = BORDER
        cell.font = Font(name="맑은 고딕", size=FS_SST, bold=True)
        cell.alignment = AC
    row += 1

    return row


THICK = Side(style="medium")   # 굵은 외곽선

def _apply_outer_border(ws, r1, c1, r2, c2):
    """
    r1~r2 행, c1~c2 열 블록의 외곽 테두리만 굵게 처리.
    내부 셀 경계는 기존 얇은 선 유지.
    """
    def _set_side(cell, side, new_side):
        b = cell.border
        cell.border = Border(
            top    = new_side if side == "top"    else b.top,
            bottom = new_side if side == "bottom" else b.bottom,
            left   = new_side if side == "left"   else b.left,
            right  = new_side if side == "right"  else b.right,
        )

    for c in range(c1, c2 + 1):
        _set_side(ws.cell(r1, c), "top",    THICK)  # 상단
        _set_side(ws.cell(r2, c), "bottom", THICK)  # 하단
    for r in range(r1, r2 + 1):
        _set_side(ws.cell(r, c1), "left",  THICK)   # 좌측
        _set_side(ws.cell(r, c2), "right", THICK)   # 우측


# ── SP 화합물 섹션 ────────────────────────────────────────────────
def _write_compound_block(ws, row, trans_list, all_trans, ref_rts=None, ref_rrt=None):
    """단일 화합물 그룹 블록 출력 (헤더 + 서브헤더 + 샘플별 데이터 행)."""
    cs = SP_DATA_COL
    show_rrt = ref_rts is not None and ref_rrt is not None
    rrt_col  = cs + N_TRANS * 2  # col 8

    # Transition 헤더 행
    for i, t in enumerate(trans_list):
        col = cs + i * 2
        _c(ws, row, col,   t["name"] + " Results", fill=BLUE_H, bold=True)
        _c(ws, row, col+1, "",                      fill=BLUE_H)
        _merge(ws, row, col, row, col + 1)
    if show_rrt:
        _c(ws, row, rrt_col,   "기준RRT", fill=BLUE_H, bold=True)
        _c(ws, row, rrt_col+1, "RRT",    fill=BLUE_H, bold=True)
        _c(ws, row, rrt_col+2, "판정",   fill=BLUE_H, bold=True)
    row += 1

    # RT / S/N 서브헤더
    for i in range(N_TRANS):
        col = cs + i * 2
        _c(ws, row, col,   "RT",  fill=BLUE_H, bold=True)
        _c(ws, row, col+1, "S/N", fill=BLUE_H, bold=True)
    if show_rrt:
        _c(ws, row, rrt_col,   "", fill=BLUE_H)
        _c(ws, row, rrt_col+1, "", fill=BLUE_H)
        _c(ws, row, rrt_col+2, "", fill=BLUE_H)
    row += 1

    # 샘플별 데이터 행
    for s_idx, sample_trans in enumerate(all_trans):
        sample_map = {t["name"]: t for t in sample_trans}
        first_rt = None
        for i, t in enumerate(trans_list):
            col = cs + i * 2
            s = sample_map.get(t["name"], {})
            _c(ws, row, col,   (round(s.get("rt"), 3) if s.get("rt") is not None else None), fill=YELLOW, num_fmt="0.000")
            _c(ws, row, col+1, (round(s.get("sn")) if s.get("sn") is not None else None), fill=YELLOW, num_fmt="0")
            if i == 0: first_rt = s.get("rt")
        if show_rrt:
            _ref = ref_rts[s_idx] if s_idx < len(ref_rts) else None
            rrt  = round(first_rt / _ref, 3) if first_rt and _ref else None
            if rrt is not None:
                ok = abs(rrt - ref_rrt) / ref_rrt <= 0.1
                rrt_fill = GREEN if ok else RED_F
                판정 = "적합" if ok else "부적합"
            else:
                rrt_fill, 판정 = GRAY, ""
            _c(ws, row, rrt_col,   ref_rrt, fill=GRAY,   num_fmt="0.0##")
            _c(ws, row, rrt_col+1, rrt,     fill=YELLOW, num_fmt="0.000")
            _c(ws, row, rrt_col+2, 판정,   fill=rrt_fill, bold=True)
        row += 1

    return row


def _write_sp_compounds(ws, row, sp_data, form_type=None, ref_rt=None):
    """SP_A / SP_B / SP_C 파일의 전체 화합물을 CSV 순서대로 출력."""
    all_trans = sp_data.get("all_transitions") or [sp_data.get("transitions", [])]
    groups = _group_transitions(all_trans[0])

    for base_name, trans_list in groups:
        while len(trans_list) < N_TRANS:
            trans_list.append({"name": f"{base_name}_?", "rt": None, "sn": None})
        _ref_rrt = RRT_REF.get(form_type, {}).get(base_name) if form_type else None
        _ref_rts = [ref_rt] * len(all_trans) if ref_rt else None
        row = _write_compound_block(ws, row, trans_list, all_trans,
                                    ref_rts=_ref_rts, ref_rrt=_ref_rrt)

    return row
