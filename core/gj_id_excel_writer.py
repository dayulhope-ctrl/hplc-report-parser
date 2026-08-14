# -*- coding: utf-8 -*-
"""공진단 확인시험 엑셀 라이터"""

import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

YELLOW  = PatternFill("solid", fgColor="FFFFE1")
BLUE_H  = PatternFill("solid", fgColor="BDD7EE")
GRAY    = PatternFill("solid", fgColor="D9D9D9")
WHITE   = PatternFill("solid", fgColor="FFFFFF")
TITLE_F = PatternFill("solid", fgColor="2E75B6")

THIN   = Side(style="thin")
MEDIUM = Side(style="medium")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left",   vertical="center")
FS = 11

COMPOUNDS = {
    "Morroniside":       (["465.3 -> 141.1", "465.3 -> 155.0", "465.3 -> 243.2"],
                          ["Morroniside (465.3 -> 243.2)", "Morroniside"]),
    "Loganin":           (["449.2 -> 101.0", "449.2 -> 127.0", "449.2 -> 227.0"],
                          ["Loganin (449.2 -> 227.0)", "Loganin"]),
    "Nodakenin":         (["409.2 -> 186.9", "409.2 -> 229.0", "409.2 -> 247.2"],
                          ["Nodakenin (409.2 -> 186.9)", "Nodakenin"]),
    "Ginsenoside Rg1":   (["859.6 -> 637.5", "859.6 -> 475.5", "859.6 -> 799.7"],
                          ["Ginsenoside Rg1 (859.6 -> 637.5)", "Ginsenoside_Rg1", "Ginsenoside Rg1"]),
    "Ginsenoside Rb1":   (["1107.6 -> 945.7", "1107.6 -> 178.8", "1107.6 -> 221.2"],
                          ["Ginsenoside Rb1 (1107.6 -> 945.7)", "Ginsenoside Rb1"]),
    "Decursin":          (["329.1 -> 247.2", "329.1 -> 115.0", "329.1 -> 128.0"],
                          ["Decursin (329.1 -> 247.2)", "Decursin"]),
    "Decursin angelate": (["329.1 -> 247.2", "329.1 -> 115.0", "329.1 -> 128.0"],
                          ["Decursin angelate (329.1 -> 247.2)", "Decursin Angelate", "Decursin angelate"]),
}

SST_GROUP1    = ["Morroniside", "Loganin", "Nodakenin"]
SST_GROUP2    = ["Ginsenoside Rg1", "Ginsenoside Rb1", "Decursin"]
RESULT_GROUP1 = ["Ginsenoside Rg1", "Ginsenoside Rb1", "Morroniside"]
RESULT_GROUP2 = ["Loganin", "Nodakenin", "Decursin"]
RESULT_GROUP3 = ["Decursin angelate"]
N_TRANS = 3


def _cell(ws, row, col, value="", fill=None, bold=False, num_fmt=None, fs=FS):
    from openpyxl.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return cell
    cell.value = value
    cell.font  = Font(name="맑은 고딕", size=fs, bold=bold)
    cell.alignment = AC
    cell.border    = BORDER
    if fill:    cell.fill = fill
    if num_fmt: cell.number_format = num_fmt
    return cell


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _title_bar(ws, row, text, c1, c2, fs=FS):
    _merge(ws, row, c1, row, c2)
    cell = ws.cell(row=row, column=c1)
    cell.value     = text
    cell.font      = Font(name="맑은 고딕", size=fs+1, bold=True, color="FFFFFF")
    cell.fill      = TITLE_F
    cell.alignment = AL
    cell.border    = BORDER
    ws.row_dimensions[row].height = 20


def _outer_border(ws, r1, c1, r2, c2):
    M = Side(style="medium")
    for c in range(c1, c2 + 1):
        b = ws.cell(r1, c).border
        ws.cell(r1, c).border = Border(top=M, bottom=b.bottom, left=b.left, right=b.right)
        b = ws.cell(r2, c).border
        ws.cell(r2, c).border = Border(top=b.top, bottom=M, left=b.left, right=b.right)
    for r in range(r1, r2 + 1):
        b = ws.cell(r, c1).border
        ws.cell(r, c1).border = Border(top=b.top, bottom=b.bottom, left=M, right=b.right)
        b = ws.cell(r, c2).border
        ws.cell(r, c2).border = Border(top=b.top, bottom=b.bottom, left=b.left, right=M)


def _lookup_trans(parsed: dict, comp_name: str, transition: str) -> dict:
    """transition 포함 키로 정확 매칭. e.g. 'Morroniside (465.3 -> 243.2)'"""
    key = f"{comp_name} ({transition})"
    if key in parsed:
        return parsed[key]
    key_lower = key.lower()
    for cname, data in parsed.items():
        if cname.lower() == key_lower:
            return data
    return {}


def _lookup_compound(parsed: dict, csv_candidates: list) -> dict:
    for name in csv_candidates:
        if name in parsed:
            return parsed[name]
    for cname in parsed:
        for cand in csv_candidates:
            if cand.lower() in cname.lower() or cname.lower() in cand.lower():
                return parsed[cname]
    return {}


def _set_widths(ws, n_comp_cols):
    ws.column_dimensions["A"].width = 16
    for ci in range(1, n_comp_cols + 1):
        ws.column_dimensions[get_column_letter(ci + 1)].width = 11


def _write_sst_group(ws, row, group_names, parsed):
    LABEL_COL = 1
    data_start_col = 2
    n_cols = len(group_names) * N_TRANS
    last_col = data_start_col + n_cols - 1

    # 헤더
    _cell(ws, row, LABEL_COL, "구분", fill=BLUE_H, bold=True)
    for gi, gname in enumerate(group_names):
        transitions, _ = COMPOUNDS[gname]
        for ti, trans in enumerate(transitions):
            col = data_start_col + gi * N_TRANS + ti
            label = f"{gname}\n({trans}) Results"
            _cell(ws, row, col, label, fill=BLUE_H, bold=True, fs=9)
    ws.row_dimensions[row].height = 35
    row += 1

    _cell(ws, row, LABEL_COL, "RT Value", fill=BLUE_H, bold=True)
    for col in range(data_start_col, last_col + 1):
        _cell(ws, row, col, "", fill=BLUE_H)
    row += 1

    # 데이터 행 1~6: transition별 RT
    data_row_start = row
    for run_idx in range(6):
        _cell(ws, row, LABEL_COL, run_idx + 1, fill=WHITE)
        for gi, gname in enumerate(group_names):
            transitions, _ = COMPOUNDS[gname]
            for ti, trans in enumerate(transitions):
                col = data_start_col + gi * N_TRANS + ti
                comp_data = _lookup_trans(parsed, gname, trans)
                rt_vals = [r.get("rt") for r in comp_data.get("std", [])]
                rt = rt_vals[run_idx] if run_idx < len(rt_vals) else None
                if rt is not None: rt = round(rt, 3)
                _cell(ws, row, col, rt, fill=YELLOW, num_fmt="0.000")
        row += 1
    data_row_end = row - 1

    # 통계 행
    for label, func in [("평균", "AVERAGE"), ("표준편차", "STDEV"), ("유지시간 RSD%\n(2.0 % 이하)", None)]:
        _cell(ws, row, LABEL_COL, label, fill=GRAY, bold=True, fs=9)
        for col in range(data_start_col, last_col + 1):
            cl = get_column_letter(col)
            rng = f"{cl}{data_row_start}:{cl}{data_row_end}"
            if func:
                formula = f"=IFERROR({func}({rng}),\"\")"
                fmt = "0.000"
            else:
                avg_row = row - 2
                std_row = row - 1
                formula = f"=IFERROR(IF({cl}{avg_row}=0,\"\",{cl}{std_row}/{cl}{avg_row}*100),\"\")"
                fmt = "0.00"
            cell = ws.cell(row, col)
            cell.value = formula
            cell.fill = GRAY
            cell.border = BORDER
            cell.font = Font(name="맑은 고딕", size=FS)
            cell.alignment = AC
            cell.number_format = fmt
        row += 1

    return row


def _write_result_group(ws, row, group_names, parsed):
    LABEL_COL = 1
    data_start_col = 2
    n_cols = len(group_names) * N_TRANS
    last_col = data_start_col + n_cols - 1

    # 헤더
    _cell(ws, row, LABEL_COL, "구분", fill=BLUE_H, bold=True)
    for gi, gname in enumerate(group_names):
        transitions, _ = COMPOUNDS[gname]
        for ti, trans in enumerate(transitions):
            col = data_start_col + gi * N_TRANS + ti
            label = f"{gname}\n({trans}) Results"
            _cell(ws, row, col, label, fill=BLUE_H, bold=True, fs=9)
    ws.row_dimensions[row].height = 35
    row += 1

    # SP 런 수 계산
    max_runs = 0
    for gname in group_names:
        transitions, _ = COMPOUNDS[gname]
        comp_data = _lookup_trans(parsed, gname, transitions[0])
        max_runs = max(max_runs, len(comp_data.get("sp", [])))
    if max_runs == 0:
        max_runs = 1

    # 첫 번째 화합물에서 SP 런 이름 목록 수집 (검체 라벨용)
    sp_run_names = []
    for gname in group_names:
        transitions, _ = COMPOUNDS[gname]
        comp_data = _lookup_trans(parsed, gname, transitions[0])
        sp_list = comp_data.get("sp", [])
        if sp_list:
            sp_run_names = [r.get("name", "") for r in sp_list]
            break

    def _short_name(name):
        import re
        n = re.sub(r'^KD_Gongjindan_', '', name)
        return n

    # S/N 데이터 행: transition별 S/N
    sn_rows = []
    for run_idx in range(max_runs):
        sn_row = row
        sn_rows.append(sn_row)
        sp_label = _short_name(sp_run_names[run_idx]) if run_idx < len(sp_run_names) else f"Run {run_idx+1}"
        _cell(ws, row, LABEL_COL, f"S/N비\n{sp_label}", fill=WHITE, bold=True, fs=9)
        ws.row_dimensions[row].height = 28
        for gi, gname in enumerate(group_names):
            transitions, _ = COMPOUNDS[gname]
            for ti, trans in enumerate(transitions):
                col = data_start_col + gi * N_TRANS + ti
                comp_data = _lookup_trans(parsed, gname, trans)
                sp_list = comp_data.get("sp", [])
                val = sp_list[run_idx].get("area") if run_idx < len(sp_list) else None
                if val is not None: val = round(val)
                _cell(ws, row, col, val, fill=YELLOW, num_fmt="0")
        row += 1

    # 결과 행: 첫 번째 transition 기준 검출/미검출
    _cell(ws, row, LABEL_COL, "결과", fill=GRAY, bold=True)
    for gi, gname in enumerate(group_names):
        col = data_start_col + gi * N_TRANS
        cl = get_column_letter(col)
        formula = f'=IFERROR(IF({cl}{sn_rows[0]}>=10,"검출","미검출"),"")' if sn_rows else ""
        _cell(ws, row, col, formula, fill=GRAY, fs=FS)
        for ti in range(1, N_TRANS):
            _cell(ws, row, data_start_col + gi * N_TRANS + ti, "", fill=GRAY)
    row += 1

    return row


def write_gj_id_result(parsed: dict) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("공진단_확인결과")
    _set_widths(ws, 9)

    row = 1

    _title_bar(ws, row, "● 시스템적합성", 1, 10)
    row += 1
    _cell(ws, row, 1, "- 표준액 6 회 조작한 결과로 확인한다.", fill=None, bold=False, fs=9)
    _merge(ws, row, 1, row, 10)
    row += 1
    _cell(ws, row, 1, "- 유지시간 RSD% 2.0% 이하이다.", fill=None, bold=False, fs=9)
    _merge(ws, row, 1, row, 10)
    row += 1

    row = _write_sst_group(ws, row, SST_GROUP1, parsed)
    row += 1
    row = _write_sst_group(ws, row, SST_GROUP2, parsed)
    row += 1

    _title_bar(ws, row, "● 결과 확인", 1, 10)
    row += 1

    row = _write_result_group(ws, row, RESULT_GROUP1, parsed)
    row += 1
    row = _write_result_group(ws, row, RESULT_GROUP2, parsed)
    row += 1
    row = _write_result_group(ws, row, RESULT_GROUP3, parsed)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
