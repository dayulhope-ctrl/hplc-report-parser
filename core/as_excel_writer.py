# -*- coding: utf-8 -*-
"""
AS 함량 결과 엑셀 생성기
맑은 고딕 14pt / 노란 셀 FFFFE1 / 양식 그대로
"""

import io, re, math
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ── 색상/스타일 상수 ──────────────────────────────────────────────
YELLOW  = PatternFill("solid", fgColor="FFFFE1")   # 입력셀
BLUE_H  = PatternFill("solid", fgColor="BDD7EE")   # 헤더
GRAY    = PatternFill("solid", fgColor="D9D9D9")   # 소계/통계
WHITE   = PatternFill("solid", fgColor="FFFFFF")
A_HEADER = PatternFill("solid", fgColor="9DC3E6")  # 검액 A 헤더 (진파랑)
B_HEADER = PatternFill("solid", fgColor="A9D18E")  # 검액 B 헤더 (진초록)
A_DATA   = PatternFill("solid", fgColor="DAEEF3")  # 검액 A 데이터셀
B_DATA   = PatternFill("solid", fgColor="E2EFDA")  # 검액 B 데이터셀

FONT_NORMAL = Font(name="맑은 고딕", size=14)
FONT_BOLD   = Font(name="맑은 고딕", size=14, bold=True)

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left",   vertical="center")

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── 헬퍼 ─────────────────────────────────────────────────────────
def _c(ws, row, col, value="", fill=None, bold=False, align="center", num_fmt=None):
    from openpyxl.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return cell   # 병합셀은 건너뜀
    cell.value = value
    cell.font   = FONT_BOLD if bold else FONT_NORMAL
    cell.alignment = ALIGN_C if align == "center" else ALIGN_L
    cell.border = BORDER
    if fill: cell.fill = fill
    if num_fmt: cell.number_format = num_fmt
    return cell

def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def _col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def _row_height(ws, row, h):
    ws.row_dimensions[row].height = h


# ── 메인 ─────────────────────────────────────────────────────────
def write_as_result(parsed: dict, lot_groups: dict) -> bytes:
    """
    parsed   : {compound: {std, sp, stability, system_check}} from as_csv_parser
    lot_groups: {lot_name: {
        "A": [{"name","area per compound"}, ...],   ← A-1, A-2
        "B": [...]                                   ← B-1, B-2
    }}
    반환: Excel bytes
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    IS_COMP  = "Cinnamic acid-d6"
    COMP_GRP1 = ["Amygdalin", "Paeoniflorin", "Cinnamic acid"]
    COMP_GRP2 = ["Ginsenoside Rb1", "Glycyrrhzin", "Decursin"]
    SP_COMPS  = ["Amygdalin", "Paeoniflorin", "Cinnamic acid",
                 "Ginsenoside Rb1", "Glycyrrhzin", "Decursin", "Decursin Angelate"]

    std_runs   = parsed.get(IS_COMP, {}).get("std", [])   # IS STD runs (6개)
    is_std_areas = [r["area"] for r in std_runs]

    if not lot_groups:
        raise ValueError("CSV에서 SP(검액) 데이터를 찾을 수 없습니다.\nCSV 파일 형식 또는 파일명을 확인해주세요.")

    for lot_name, lot_data in lot_groups.items():
        ws = wb.create_sheet(title=lot_name[:31])
        _build_sheet(ws, parsed, is_std_areas, lot_data,
                     IS_COMP, COMP_GRP1, COMP_GRP2, SP_COMPS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_sheet(ws, parsed, is_std_areas, lot_data,
                 IS_COMP, COMP_GRP1, COMP_GRP2, SP_COMPS):

    # 열 너비 설정 (STD/IS 제거 후: 구분1 + 화합물당 2열 × 3 = 7열)
    _col_widths(ws, {
        1: 12,   # A: 구분
        2: 10,   # B: RT
        3: 14,   # C: Response
        4: 10,   # D: RT
        5: 14,   # E: Response
        6: 10,   # F: RT
        7: 14,   # G: Response
        8: 20,   # H: SP 화합물명 (긴 이름 대응)
        9: 20,   # I
        10: 20,  # J
        11: 20,  # K
        12: 20,  # L
        13: 20,  # M
        14: 20,  # N
    })

    row = 1  # 현재 행 포인터

    # ══════════════════════════════════════════════════════════
    # 4) 표준액 정보
    # ══════════════════════════════════════════════════════════
    _c(ws, row, 1, "4) 표준액 정보", bold=True)
    _merge(ws, row, 1, row, 14)
    _row_height(ws, row, 22)
    row += 1

    # ── IS 표 ─────────────────────────────────────────────────
    row = _is_table(ws, row, parsed.get(IS_COMP, {}), IS_COMP)

    row += 1  # 빈 행

    # ── 화합물 그룹1 (Amygdalin, Paeoniflorin, Cinnamic acid) ──
    row = _compound_group_table(ws, row, COMP_GRP1, parsed, is_std_areas)

    row += 1

    # ── 화합물 그룹2 (Ginsenoside Rb1, Glycyrrhzin, Decursin) ──
    row = _compound_group_table(ws, row, COMP_GRP2, parsed, is_std_areas)

    row += 1

    # ══════════════════════════════════════════════════════════
    # 5) 검액 정보
    # ══════════════════════════════════════════════════════════
    _c(ws, row, 1, "5) 검액 정보", bold=True)
    _merge(ws, row, 1, row, 14)
    _row_height(ws, row, 22)
    row += 1

    row = _sp_is_table(ws, row, IS_COMP, lot_data)
    row += 1
    row = _sp_compound_table(ws, row, SP_COMPS, lot_data)

    row += 1


# ── IS 표 ────────────────────────────────────────────────────────
def _is_table(ws, row, is_data, comp_name):
    std_rows = is_data.get("std", [])

    # 헤더
    _c(ws, row, 1, "구분",    fill=BLUE_H, bold=True)
    _c(ws, row, 2, comp_name, fill=BLUE_H, bold=True)
    _merge(ws, row, 2, row, 3)
    # 병합 후 merged cell에는 쓰지 않음
    row += 1

    _c(ws, row, 1, "",          fill=BLUE_H)
    _c(ws, row, 2, "RT",        fill=BLUE_H, bold=True)
    _c(ws, row, 3, "Response",  fill=BLUE_H, bold=True)
    row += 1

    # Value 행 6개
    val_start = row
    for i in range(6):
        r = std_rows[i] if i < len(std_rows) else {}
        if i == 0:
            _c(ws, row, 1, "Value", fill=WHITE, bold=True)
        else:
            _c(ws, row, 1, "", fill=WHITE)
        _c(ws, row, 2, r.get("rt"),   fill=YELLOW, num_fmt="0.000")
        _c(ws, row, 3, r.get("area"), fill=YELLOW, num_fmt="0")
        row += 1
    if val_start < row - 1:
        _merge(ws, val_start, 1, row - 1, 1)

    # 평균
    r2, r3 = f"B{val_start}", f"C{val_start}"
    r2e, r3e = f"B{row-1}", f"C{row-1}"
    _c(ws, row, 1, "평균", fill=GRAY, bold=True)
    ws.cell(row, 2, f'=IFERROR(AVERAGE(B{val_start}:B{row-1}),"")').number_format = "0.000"
    ws.cell(row, 2).fill = GRAY; ws.cell(row, 2).border = BORDER
    ws.cell(row, 2).font = FONT_NORMAL; ws.cell(row, 2).alignment = ALIGN_C
    ws.cell(row, 3, f'=IFERROR(AVERAGE(C{val_start}:C{row-1}),"")').number_format = "0"
    ws.cell(row, 3).fill = GRAY; ws.cell(row, 3).border = BORDER
    ws.cell(row, 3).font = FONT_NORMAL; ws.cell(row, 3).alignment = ALIGN_C
    row += 1

    return row


# ── 화합물 그룹 표 ────────────────────────────────────────────────
def _compound_group_table(ws, row, compounds, parsed, is_std_areas):
    """3개 화합물을 가로로 나란히 배치 (RT + Response 2열씩)"""
    # 구분 열 = col 1
    # comp1: 2,3 / comp2: 4,5 / comp3: 6,7
    COL_STARTS = [2, 4, 6]

    # 화합물 헤더
    _c(ws, row, 1, "구분", fill=BLUE_H, bold=True)
    for i, comp in enumerate(compounds):
        cs = COL_STARTS[i]
        _c(ws, row, cs,   comp, fill=BLUE_H, bold=True)
        _c(ws, row, cs+1, "",   fill=BLUE_H)
        _merge(ws, row, cs, row, cs+1)
    row += 1

    # 서브헤더
    _c(ws, row, 1, "", fill=BLUE_H)
    for cs in COL_STARTS:
        _c(ws, row, cs,   "RT",       fill=BLUE_H, bold=True)
        _c(ws, row, cs+1, "Response", fill=BLUE_H, bold=True)
    row += 1

    # Value 행 6개
    val_start = row
    for i in range(6):
        if i == 0:
            _c(ws, row, 1, "Value", fill=WHITE, bold=True)
        else:
            _c(ws, row, 1, "", fill=WHITE)

        for j, comp in enumerate(compounds):
            cs = COL_STARTS[j]
            comp_data = parsed.get(comp, {}).get("std", [])
            r = comp_data[i] if i < len(comp_data) else {}
            _c(ws, row, cs,   r.get("rt"),   fill=YELLOW, num_fmt="0.000")
            _c(ws, row, cs+1, r.get("area"), fill=YELLOW, num_fmt="0")

        row += 1

    if val_start < row - 1:
        _merge(ws, val_start, 1, row - 1, 1)

    # 통계 행 (평균, 표준편차, %RSD)
    for stat_label, stat_func in [("평균", "AVERAGE"), ("표준편차", "STDEV"), ("%RSD", None)]:
        _c(ws, row, 1, stat_label, fill=GRAY, bold=True)
        for cs in COL_STARTS:
            resp_col = get_column_letter(cs + 1)
            rng = f"{resp_col}{val_start}:{resp_col}{val_start+5}"
            if stat_func:
                formula = f'=IFERROR({stat_func}({rng}),"")'
            else:
                avg_row = row - 2
                std_row = row - 1
                formula = f'=IFERROR(IF({resp_col}{avg_row}=0,"",{resp_col}{std_row}/{resp_col}{avg_row}*100),"")'
            cell = ws.cell(row, cs + 1, formula)
            cell.fill = GRAY; cell.border = BORDER
            cell.font = FONT_NORMAL; cell.alignment = ALIGN_C
            cell.number_format = "0.00"
            if stat_label == "평균":
                rt_col = get_column_letter(cs)
                cell = ws.cell(row, cs, f'=IFERROR(AVERAGE({rt_col}{val_start}:{rt_col}{val_start+5}),"")' )
                cell.fill = GRAY; cell.border = BORDER
                cell.font = FONT_NORMAL; cell.alignment = ALIGN_C
                cell.number_format = "0.000"
            else:
                _c(ws, row, cs, fill=GRAY)
        row += 1

    return row


# ── SP IS 표 ─────────────────────────────────────────────────────
def _sp_is_table(ws, row, is_comp, lot_data):
    a_runs = lot_data.get("A", [])
    b_runs = lot_data.get("B", [])

    _c(ws, row, 1, "구분\n(I.S)", fill=BLUE_H, bold=True)
    _c(ws, row, 2, f"{is_comp}\n검액 A", fill=A_HEADER, bold=True)
    _c(ws, row, 3, f"{is_comp}\n검액 B", fill=B_HEADER, bold=True)
    ws.row_dimensions[row].height = 36
    row += 1

    _c(ws, row, 1, "",         fill=BLUE_H)
    _c(ws, row, 2, "Response", fill=A_HEADER, bold=True)
    _c(ws, row, 3, "Response", fill=B_HEADER, bold=True)
    row += 1

    for sp_idx, label in enumerate(["SP1", "SP2"]):
        a_area = a_runs[sp_idx].get(is_comp) if sp_idx < len(a_runs) else None
        b_area = b_runs[sp_idx].get(is_comp) if sp_idx < len(b_runs) else None
        _c(ws, row, 1, label, fill=BLUE_H, bold=True)
        _c(ws, row, 2, a_area, fill=YELLOW, num_fmt="0")
        _c(ws, row, 3, b_area, fill=YELLOW, num_fmt="0")
        row += 1

    return row


# ── SP 화합물 표 ─────────────────────────────────────────────────
def _sp_compound_table(ws, row, sp_comps, lot_data):
    a_runs = lot_data.get("A", [])
    b_runs = lot_data.get("B", [])

    A_COMPS = {"Cinnamic acid", "Ginsenoside Rb1"}
    B_COMPS = {"Amygdalin", "Paeoniflorin", "Glycyrrhzin",
               "Decursin", "Decursin Angelate"}

    def _comp_hdr(comp):
        return A_HEADER if comp in A_COMPS else B_HEADER

    def _comp_data(comp):
        return A_DATA if comp in A_COMPS else B_DATA

    # 화합물 헤더
    _c(ws, row, 1, "구분", fill=BLUE_H, bold=True)
    for ci, comp in enumerate(sp_comps, 2):
        _c(ws, row, ci, comp, fill=_comp_hdr(comp), bold=True)
    ws.row_dimensions[row].height = 40
    row += 1

    # 서브헤더
    _c(ws, row, 1, "", fill=BLUE_H)
    for ci, comp in enumerate(sp_comps, 2):
        _c(ws, row, ci, "Response", fill=_comp_hdr(comp), bold=True)
    row += 1

    # 데이터 행
    for sp_idx, label in enumerate(["SP1", "SP2"]):
        _c(ws, row, 1, label, fill=BLUE_H, bold=True)
        for ci, comp in enumerate(sp_comps, 2):
            runs = a_runs if comp in A_COMPS else b_runs
            area = runs[sp_idx].get(comp) if sp_idx < len(runs) else None
            _c(ws, row, ci, area, fill=YELLOW, num_fmt="0")
        row += 1

    return row


# ── 계산 표 ──────────────────────────────────────────────────────
CALC_COMPS = [
    ("Amygdalin",       "행인 중 아미그달린",        "mg/1환"),
    ("Paeoniflorin",    "작약 중 패오니플로린",       "mg/1환"),
    ("Cinnamic acid",   "육계 중 신남산",            "μg/1환"),
    ("Ginsenoside Rb1", "인삼 중 진세노시드 Rb1",    "mg/1환"),
    ("Glycyrrhzin",     "감초 중 글리시리진산",       "mg/1환"),
    ("Decursin",        "당귀 중 총데쿠르신",         "mg/1환"),
]

def _calc_table(ws, row):
    _c(ws, row, 1, "구분",     fill=BLUE_H, bold=True)
    _c(ws, row, 2, "Sample 1", fill=BLUE_H, bold=True)
    _merge(ws, row, 2, row, 3)
    _c(ws, row, 3, "", fill=BLUE_H)
    _c(ws, row, 4, "Sample 2", fill=BLUE_H, bold=True)
    _merge(ws, row, 4, row, 5)
    _c(ws, row, 5, "", fill=BLUE_H)
    _c(ws, row, 6, "Avg.",     fill=BLUE_H, bold=True)
    _merge(ws, row, 6, row, 7)
    _c(ws, row, 7, "", fill=BLUE_H)
    row += 1

    for comp, korean, unit in CALC_COMPS:
        _c(ws, row, 1, korean, fill=WHITE, bold=True, align="left")
        _c(ws, row, 2, "",    fill=YELLOW)   # Sample 1 계산값 (수동입력 또는 추후 수식)
        _c(ws, row, 3, unit,  fill=WHITE)
        _c(ws, row, 4, "",    fill=YELLOW)   # Sample 2
        _c(ws, row, 5, unit,  fill=WHITE)
        _c(ws, row, 6, "",    fill=GRAY)     # Avg (수식)
        _c(ws, row, 7, unit,  fill=WHITE)
        row += 1

    return row
