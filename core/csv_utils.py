# -*- coding: utf-8 -*-
"""
CSV 파싱 공통 유틸리티

모든 MassHunter CSV/Excel 파일에 공통으로 사용하는 자동 파서입니다.
- NAME_COL  : Row 1에서 "Name" 열 위치 자동 탐색
- 화합물 컬럼: Row 0에서 "X Results" 위치 및 Row 1의 RT/Area(S/N) 컬럼 확인
인식되지 않은 모든 키워드(Results, Name, RT, Area, S/N)는 무시됩니다.
"""

import csv, re, math, io
from collections import defaultdict


# ── 기본 I/O ──────────────────────────────────────────────────────────
def _is_xlsx(file_obj) -> bool:
    try:
        if hasattr(file_obj, "read"):
            header = file_obj.read(4)
            if hasattr(file_obj, "seek"):
                file_obj.seek(0)
            return header[:2] == b"PK"
        else:
            with open(file_obj, "rb") as f:
                return f.read(2) == b"PK"
    except Exception:
        return False


def _read_xlsx_rows(file_obj) -> list:
    import openpyxl
    if hasattr(file_obj, "read"):
        data = file_obj.read()
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            elif isinstance(v, float):
                cells.append(repr(v))
            else:
                cells.append(str(v))
        rows.append(cells)
    wb.close()
    return rows


def read_rows(file_obj) -> list:
    if _is_xlsx(file_obj):
        return _read_xlsx_rows(file_obj)
    if hasattr(file_obj, "read"):
        content = file_obj.read()
        if isinstance(content, bytes):
            content = content.decode("utf-8-sig", errors="replace")
        lines = content.splitlines()
    else:
        with open(file_obj, encoding="utf-8-sig", errors="replace") as f:
            lines = f.read().splitlines()
    return list(csv.reader(lines))


def safe_float(val):
    if isinstance(val, (int, float)):
        return float(val)
    try:
        v = val.strip()
        return float(v) if v else None
    except (ValueError, AttributeError):
        return None


# ── 헤더 자동 감지 ────────────────────────────────────────────────────
def detect_columns(rows: list, prefer: str = "area") -> tuple:
    """
    Row 0 (화합물 명), Row 1 (서브헤더)을 스캔해
    name_col 및 {compound: (rt_col, value_col)} 를 반환.

    prefer="area"  -> Area 우선, 없으면 S/N (AS 입력용)
    prefer="sn"    -> S/N 우선, 없으면 Area (ID 확인용)
    """
    if len(rows) < 2:
        return 0, {}

    header0 = rows[0]
    header1 = rows[1]

    # NAME_COL: row 1에서 "name" 텍스트 위치
    name_col = 0
    for ci, val in enumerate(header1):
        if val.strip().lower() == "name":
            name_col = ci
            break

    # 화합물 컬럼: row 0에서 "Results" 포함 셀 탐색
    compound_cols = {}
    for ci, val in enumerate(header0):
        if "Results" not in val:
            continue
        comp = val.replace("Results", "").strip()
        if not comp:
            continue

        rt_col = area_col = sn_col = None
        for offset in range(5):
            idx = ci + offset
            if idx >= len(header1):
                break
            sub = header1[idx].strip().upper()
            if sub == "RT" and rt_col is None:
                rt_col = idx
            elif sub == "AREA" and area_col is None:
                area_col = idx
            elif sub == "S/N" and sn_col is None:
                sn_col = idx

        if rt_col is None:
            continue

        # prefer에 따라 value 컬럼 선택
        if prefer == "area":
            val_col = area_col if area_col is not None else (sn_col if sn_col is not None else rt_col + 1)
        else:  # "sn"
            val_col = sn_col if sn_col is not None else (area_col if area_col is not None else rt_col + 1)

        compound_cols[comp] = (rt_col, val_col)

    return name_col, compound_cols


# ── 런 분류 ────────────────────────────────────────────────────────────
def classify(name: str) -> str:
    n = name.upper()
    if "SYSTEM" in n:     return "system_check"
    if "STABILITY" in n:  return "stability"
    if "STD" in n:        return "std"
    return "sp"


# ── 단일 파싱 ──────────────────────────────────────────────────────────
def parse_csv(file_obj, prefer: str = "area") -> dict:
    """
    단일 파싱. 반환:
    {
      compound: {
        "std":          [{"name","rt","area"}, ...],
        "sp":           [...],
        "stability":    [...],
        "system_check": [...],
      }
    }
    """
    rows = read_rows(file_obj)
    if len(rows) < 3:
        return {}

    name_col, compound_cols = detect_columns(rows, prefer=prefer)
    if not compound_cols:
        return {}

    result = defaultdict(lambda: {"std": [], "sp": [], "stability": [], "system_check": []})

    for row in rows[2:]:
        if len(row) < 2:
            continue
        name = row[name_col].strip() if name_col < len(row) else ""
        if not name:
            continue

        section = classify(name)

        for comp, (rt_col, val_col) in compound_cols.items():
            rt   = safe_float(row[rt_col])  if rt_col  < len(row) else None
            area = safe_float(row[val_col]) if val_col < len(row) else None
            if rt is None and area is None:
                continue
            result[comp][section].append({"name": name, "rt": rt, "area": area})

    return dict(result)


# ── SP 런 lot 그룹화 ────────────────────────────────────────────────────
def group_by_lot(parsed: dict) -> dict:
    """
    SP 런을 lot명·A/B 메서드별로 그룹화.
    이름 예시: <LotName>_<A|B>-<num>  (예: BSHwan_26006_B-1)
    반환: {lot_name: {"A": [{comp: area, ...}, ...], "B": [...]}}
    """
    run_map = {}
    for comp, data in parsed.items():
        for entry in data.get("sp", []):
            name = entry["name"]
            if name not in run_map:
                run_map[name] = {"name": name}
            run_map[name][comp] = entry.get("area")

    lots = {}
    for run_name, run_data in run_map.items():
        m = re.search(r'^(.+?)_([AB])-(\d+)$', run_name)
        if not m:
            continue
        lot_name = m.group(1)
        method   = m.group(2)
        run_num  = int(m.group(3))

        if lot_name not in lots:
            lots[lot_name] = {"A": [], "B": []}

        existing = lots[lot_name][method]
        while len(existing) < run_num:
            existing.append({})
        existing[run_num - 1] = run_data

    return lots


# ── 복수 파일 병합 ────────────────────────────────────────────────────
def merge_parsed(file_list, prefer: str = "area") -> dict:
    """
    복수 CSV/Excel 파일(또는 파일 객체 리스트)을 모두 파싱한 뒤 결과를 합침.
    """
    merged = defaultdict(lambda: {"std": [], "sp": [], "stability": [], "system_check": []})
    for f in file_list:
        partial = parse_csv(f, prefer=prefer)
        for comp, sections in partial.items():
            for sec, rows in sections.items():
                merged[comp][sec].extend(rows)
    return dict(merged)


# ── STD 통계 계산 ────────────────────────────────────────────────────
def get_stats(compound_data: dict) -> dict:
    """STD 런 통계 (평균·표준편차·%RSD)"""
    std_rows = compound_data.get("std", [])
    if not std_rows:
        return {}
    areas = [r["area"] for r in std_rows if r["area"] is not None]
    rts   = [r["rt"]   for r in std_rows if r["rt"]   is not None]
    if not areas:
        return {}
    avg = sum(areas) / len(areas)
    sd  = math.sqrt(sum((x - avg) ** 2 for x in areas) / (len(areas) - 1)) if len(areas) > 1 else 0
    rsd = round(sd / avg * 100, 3) if avg else 0
    return {
        "count":     len(areas),
        "area_list": areas,
        "rt_list":   rts,
        "avg_area":  round(avg, 1),
        "avg_rt":    round(sum(rts) / len(rts), 3) if rts else None,
        "sd":        round(sd, 2),
        "rsd":       rsd,
    }
