# -*- coding: utf-8 -*-
"""
아미노산 함량 분석 엑셀 라이터
시트 구성:
  1) 시스템적합성  — 나눔고딕 14 / 분리도 6런 + 판정
  2) 표준액 면적   — 맑은 고딕 14 / 6런 면적, 평균, %RSD + 판정  (AA당 2열)
  3) 검액_{lot}   — 맑은 고딕 14 / Sample 1/2 면적             (AA당 2열)
"""
import io
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.aa_pdf_parser import AA_ORDER

# ── 폰트 ──────────────────────────────────────────────────────────
FN_SST  = "나눔고딕"    # 시스템적합성
FN_DATA = "맑은 고딕"  # 표준액 면적 / 검액 면적
FS      = 14

# ── 채우기 ────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor="BDD7EE")   # 파란 헤더
FILL_GRAY   = PatternFill("solid", fgColor="D9D9D9")   # 행 레이블
FILL_DATA   = PatternFill("solid", fgColor="FFFFE1")   # 데이터 셀 (연노랑)
FILL_PASS   = PatternFill("solid", fgColor="C6EFCE")   # 적합 (초록)
FILL_FAIL   = PatternFill("solid", fgColor="FFC7CE")   # 부적합 (빨강)

# ── 테두리 ────────────────────────────────────────────────────────
_T = Side(style="thin")
_BORDER = Border(left=_T, right=_T, top=_T, bottom=_T)


def _c(ws, row, col, value=None, *,
       font_name=FN_DATA, fill=None, bold=False,
       align="center", border=True):
    """셀 하나 쓰기 헬퍼."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=font_name, size=FS, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if border:
        c.border = _BORDER
    if fill:
        c.fill = fill
    return c


_mc_style_cache: dict = {}   # fill_key → MergedCell용 _style 인덱스


def _mc_right_style(ws, fill):
    """MergedCell에 적용할 right+top+bottom border _style 인덱스 반환 (캐시)."""
    key = fill.fgColor.rgb if fill else "none"
    if key not in _mc_style_cache:
        tmp_r, tmp_c = 9990 + len(_mc_style_cache), 9999
        tmp = ws.cell(row=tmp_r, column=tmp_c)
        tmp.border = Border(right=_T, top=_T, bottom=_T)
        if fill:
            tmp.fill = fill
        _mc_style_cache[key] = tmp._style
        ws._cells.pop((tmp_r, tmp_c), None)   # 임시 셀 제거
    return _mc_style_cache[key]


def _twin(ws, row, c1, c2, value, fn, fill=None, bold=False):
    """
    c1:c2 병합 + 완전한 테두리.
    - 마스터 셀(c1): left·top·bottom border + fill + value
    - MergedCell(c2): _style 직접 설정으로 right·top·bottom border + fill 적용
    """
    c1_letter = get_column_letter(c1)
    c2_letter = get_column_letter(c2)
    ws.merge_cells(f"{c1_letter}{row}:{c2_letter}{row}")

    tl = ws.cell(row=row, column=c1, value=value)
    tl.font      = Font(name=fn, size=FS, bold=bold)
    tl.alignment = Alignment(horizontal="center", vertical="center")
    tl.border    = Border(left=_T, top=_T, bottom=_T)
    if fill:
        tl.fill = fill

    mc = ws._cells.get((row, c2))
    if mc is not None:
        mc._style = _mc_right_style(ws, fill)

    return tl


def _calc_rsd(values: list):
    if len(values) < 2:
        return None
    avg = sum(values) / len(values)
    if avg == 0:
        return None
    sd = math.sqrt(sum((x - avg) ** 2 for x in values) / (len(values) - 1))
    return sd / avg * 100


# ══════════════════════════════════════════════════════════════════
# 시트 1 : 시스템적합성  (나눔고딕 14)
# ══════════════════════════════════════════════════════════════════
def _sheet_sst(wb, resolutions: list[dict]):
    ws = wb.create_sheet("시스템적합성")

    ws.column_dimensions["A"].width = 12
    for i in range(2, 10):
        ws.column_dimensions[get_column_letter(i)].width = 9

    def cv(row, col, val=None, *, fill=None, bold=False):
        return _c(ws, row, col, val,
                  font_name=FN_SST, fill=fill, bold=bold)

    # 헤더
    cv(1, 1, "분리도", fill=FILL_HEADER, bold=True)
    for i in range(1, 7):
        cv(1, i + 1, str(i), fill=FILL_HEADER, bold=True)
    cv(1, 8, "판정", fill=FILL_HEADER, bold=True)

    for row_i, aa in enumerate(AA_ORDER):
        r = row_i + 2
        cv(r, 1, aa, fill=FILL_DATA)

        res_vals = [
            resolutions[ri].get(aa) if ri < len(resolutions) else None
            for ri in range(6)
        ]

        for col_i, rv in enumerate(res_vals):
            cv(r, col_i + 2,
               round(rv, 3) if rv is not None else "N/A",
               fill=FILL_DATA)

        valid = [v for v in res_vals if v is not None]
        if not valid:
            cv(r, 8, "N/A", fill=FILL_DATA)
        elif all(v >= 1.2 for v in valid):
            cv(r, 8, "적합",   fill=FILL_PASS, bold=True)
        else:
            cv(r, 8, "부적합", fill=FILL_FAIL, bold=True)

    ws.row_dimensions[1].height = 20
    return ws


# ══════════════════════════════════════════════════════════════════
# 시트 2 : 표준액 면적  (맑은 고딕 14 / AA당 2열 전체 병합)
# 모든 행(헤더·런·평균·%RSD)에서 AA당 c1:c1+1 병합
# 판정 열 없음 — %RSD 셀 색상(초록/빨강)으로 적합 여부 표시
# ══════════════════════════════════════════════════════════════════
def _write_std_group(ws, start_row: int, group_aas: list, runs: list) -> int:
    """그룹 하나 작성. 다음 시작 행 반환."""

    _c(ws, start_row, 1, "", fill=FILL_HEADER)

    for i, aa in enumerate(group_aas):
        c1 = 2 + i * 2
        _twin(ws, start_row, c1, c1 + 1, aa, FN_DATA, fill=FILL_HEADER, bold=True)

    for ri, run_data in enumerate(runs[:6]):
        r = start_row + 1 + ri
        _c(ws, r, 1, str(ri + 1), fill=FILL_GRAY, bold=True)
        for i, aa in enumerate(group_aas):
            c1 = 2 + i * 2
            val = run_data.get(aa)
            _twin(ws, r, c1, c1 + 1,
                   round(val, 3) if val is not None else "", FN_DATA, fill=FILL_DATA)

    avg_r = start_row + 7
    _c(ws, avg_r, 1, "평균", fill=FILL_GRAY, bold=True)
    for i, aa in enumerate(group_aas):
        c1 = 2 + i * 2
        vals = [run[aa] for run in runs if aa in run]
        avg  = round(sum(vals) / len(vals), 3) if vals else ""
        _twin(ws, avg_r, c1, c1 + 1, avg, FN_DATA, fill=FILL_DATA)

    rsd_r = start_row + 8
    _c(ws, rsd_r, 1, "%RSD", fill=FILL_GRAY, bold=True)
    for i, aa in enumerate(group_aas):
        c1 = 2 + i * 2
        vals = [run[aa] for run in runs if aa in run]
        rsd  = _calc_rsd(vals)
        if rsd is None:
            rsd_fill = FILL_DATA
        elif rsd <= 2.0:
            rsd_fill = FILL_PASS
        else:
            rsd_fill = FILL_FAIL
        _twin(ws, rsd_r, c1, c1 + 1,
               round(rsd, 1) if rsd is not None else "", FN_DATA,
               fill=rsd_fill, bold=(rsd is not None))

    return rsd_r + 2


def _sheet_std_area(wb, runs: list[dict]):
    ws = wb.create_sheet("표준액 면적")

    ws.column_dimensions["A"].width = 8
    for i in range(len(AA_ORDER)):
        # 두 열 합산 너비를 값 열 하나에 설정 (판정 열 제거로 시각 보정)
        ws.column_dimensions[get_column_letter(2 + i * 2)    ].width = 14
        ws.column_dimensions[get_column_letter(2 + i * 2 + 1)].width = 7

    groups = [AA_ORDER[:5], AA_ORDER[5:10], AA_ORDER[10:]]
    cur = 1
    for grp in groups:
        cur = _write_std_group(ws, cur, grp, runs)

    return ws


# ══════════════════════════════════════════════════════════════════
# 시트 3+ : 검액 면적  (맑은 고딕 14 / AA당 2열 전체 병합)
# ══════════════════════════════════════════════════════════════════
def _write_sp_group(ws, start_row: int, group_aas: list, lot_data: dict) -> int:
    """검액 그룹 하나 작성. 다음 시작 행 반환."""

    _c(ws, start_row, 1, "", fill=FILL_HEADER)

    for i, aa in enumerate(group_aas):
        c1 = 2 + i * 2
        _twin(ws, start_row, c1, c1 + 1, aa, FN_DATA, fill=FILL_HEADER, bold=True)

    for s in (1, 2):
        r = start_row + s
        _c(ws, r, 1, f"Sample {s}", fill=FILL_GRAY, bold=True)
        for i, aa in enumerate(group_aas):
            c1 = 2 + i * 2
            val = lot_data.get(s, {}).get(aa)
            _twin(ws, r, c1, c1 + 1,
                   round(val, 3) if val is not None else "", FN_DATA, fill=FILL_DATA)

    return start_row + 4


def _sheet_sp(wb, lot_id: str, lot_data: dict):
    ws = wb.create_sheet(f"검액_{lot_id}"[:31])

    ws.column_dimensions["A"].width = 12
    for i in range(len(AA_ORDER)):
        ws.column_dimensions[get_column_letter(2 + i * 2)    ].width = 14
        ws.column_dimensions[get_column_letter(2 + i * 2 + 1)].width = 7

    groups = [AA_ORDER[:5], AA_ORDER[5:10], AA_ORDER[10:]]
    cur = 1
    for grp in groups:
        cur = _write_sp_group(ws, cur, grp, lot_data)

    return ws


# ══════════════════════════════════════════════════════════════════
# 진입점
# ══════════════════════════════════════════════════════════════════
def write_aa_result(runs: list, resolutions: list, lots: dict) -> bytes:
    global _mc_style_cache
    _mc_style_cache = {}   # 워크북마다 스타일 인덱스 초기화

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_sst(wb, resolutions)
    _sheet_std_area(wb, runs)
    for lot_id in sorted(lots.keys()):
        _sheet_sp(wb, lot_id, lots[lot_id])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
