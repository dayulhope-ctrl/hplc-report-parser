# -*- coding: utf-8 -*-
"""결과 엑셀 생성"""

import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

YELLOW = PatternFill("solid", fgColor="FFFF00")
GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
HEADER = PatternFill("solid", fgColor="D9D9D9")

CENTER = Alignment(horizontal="center", vertical="center")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def _cell(ws, row, col, value, fill=None, bold=False, align=True):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    if bold:
        c.font = Font(bold=True)
    if align:
        c.alignment = CENTER
    c.border = BORDER
    return c


def write_id_result(result_rows: list, lot_info: str = "") -> bytes:
    """
    확인 결과 엑셀 생성
    result_rows: [
      {
        "compound": "Platycodin-D",
        "transitions": ["1223.7","681.1","469.2"],
        "rrt_expected": 0.8,
        "samples": {
          "26006_A": {"rrt_calc": 0.82, "sn": 170.7, "result": "적합"},
          ...
        }
      }, ...
    ]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "확인 결과"

    sample_names = []
    if result_rows:
        sample_names = list(result_rows[0]["samples"].keys())

    # 헤더
    col = 1
    _cell(ws, 1, col, "성분", fill=HEADER, bold=True); col += 1
    _cell(ws, 1, col, "전이", fill=HEADER, bold=True); col += 1
    _cell(ws, 1, col, "RRT 기준", fill=HEADER, bold=True); col += 1
    for sn in sample_names:
        _cell(ws, 1, col, f"{sn} RT", fill=HEADER, bold=True); col += 1
        _cell(ws, 1, col, f"{sn} S/N", fill=HEADER, bold=True); col += 1
        _cell(ws, 1, col, f"{sn} RRT", fill=HEADER, bold=True); col += 1
        _cell(ws, 1, col, f"{sn} 결과", fill=HEADER, bold=True); col += 1

    # 데이터
    for r_idx, row in enumerate(result_rows, start=2):
        col = 1
        _cell(ws, r_idx, col, row["compound"]); col += 1
        _cell(ws, r_idx, col, " / ".join(row["transitions"])); col += 1
        _cell(ws, r_idx, col, row["rrt_expected"]); col += 1
        for sn in sample_names:
            s = row["samples"].get(sn, {})
            _cell(ws, r_idx, col, s.get("rt")); col += 1
            _cell(ws, r_idx, col, s.get("sn")); col += 1
            _cell(ws, r_idx, col, s.get("rrt_calc")); col += 1
            result_val = s.get("result", "")
            fill = GREEN if result_val == "적합" else (RED if result_val == "부적합" else None)
            _cell(ws, r_idx, col, result_val, fill=fill); col += 1

    # 열 너비
    for i in range(1, col):
        ws.column_dimensions[get_column_letter(i)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_as_result(result_data: dict) -> bytes:
    """
    함량 결과 엑셀 생성
    result_data: {
      "std_info": {compound: {rt_list, resp_list, is_ratio_list, avg_rt, avg_resp, avg_ratio, sd, rsd}},
      "is_info": {"rt_avg": ..., "resp_avg": ..., "resp_list": [...]},
      "sp_info": {lot: {compound: {sp_resp, is_resp, content}}},
      "params": {compound: {korean, unit, ...}}
    }
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "함량 결과"

    row = 1

    # IS 표준액 정보
    is_info = result_data.get("is_info", {})
    _cell(ws, row, 1, "4) 표준액 정보 - IS (Cinnamic acid-d6)", fill=HEADER, bold=True)
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    for hdr, val in [("평균 RT", is_info.get("rt_avg")), ("평균 Response", is_info.get("resp_avg"))]:
        _cell(ws, row, 1, hdr, fill=HEADER, bold=True)
        _cell(ws, row, 2, val)
        row += 1
    row += 1

    # 화합물별 표준액 정보
    std_info = result_data.get("std_info", {})
    params = result_data.get("params", {})

    for comp, sdata in std_info.items():
        korean = params.get(comp, {}).get("korean", comp)
        unit = params.get(comp, {}).get("unit", "mg")
        _cell(ws, row, 1, f"{korean} 표준액", fill=HEADER, bold=True)
        ws.merge_cells(f"A{row}:H{row}")
        row += 1

        headers = ["구분", "RT", "Response", "STD/I.S"]
        for ci, h in enumerate(headers, 1):
            _cell(ws, row, ci, h, fill=HEADER, bold=True)
        row += 1

        resp_list = sdata.get("resp_list", [])
        rt_list   = sdata.get("rt_list", [])
        ratio_list = sdata.get("is_ratio_list", [])
        for i in range(len(resp_list)):
            _cell(ws, row, 1, f"STD-{i+1}")
            _cell(ws, row, 2, rt_list[i] if i < len(rt_list) else "")
            _cell(ws, row, 3, resp_list[i])
            _cell(ws, row, 4, ratio_list[i] if i < len(ratio_list) else "")
            row += 1

        for label, key in [("평균", "avg_resp"), ("표준편차", "sd"), ("%RSD", "rsd")]:
            _cell(ws, row, 1, label, bold=True)
            _cell(ws, row, 3, sdata.get(key))
            if key == "avg_resp":
                _cell(ws, row, 4, sdata.get("avg_ratio"))
            row += 1
        row += 1

    # 검액 함량 결과
    sp_info = result_data.get("sp_info", {})
    _cell(ws, row, 1, "5) 검액 함량 결과", fill=HEADER, bold=True)
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    compounds = list(std_info.keys())
    headers = ["Lot"] + [f"{params.get(c,{}).get('korean',c)}\n({params.get(c,{}).get('unit','mg')}/환)" for c in compounds]
    for ci, h in enumerate(headers, 1):
        _cell(ws, row, ci, h, fill=HEADER, bold=True)
    row += 1

    for lot, lot_data in sp_info.items():
        _cell(ws, row, 1, lot)
        for ci, comp in enumerate(compounds, 2):
            val = lot_data.get(comp, {}).get("content")
            _cell(ws, row, ci, val if val is not None else "N/A")
        row += 1

    for i in range(1, len(headers) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
