# -*- coding: utf-8 -*-
"""공진단 함량 엑셀 라이터 — 템플릿 없이 생성"""
import io, re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

YELLOW = PatternFill("solid", fgColor="FFFFE1")
BLUE_H = PatternFill("solid", fgColor="BDD7EE")
GRAY   = PatternFill("solid", fgColor="D9D9D9")
WHITE  = PatternFill("solid", fgColor="FFFFFF")
TITLE_F= PatternFill("solid", fgColor="2E75B6")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left",   vertical="center")
FS = 12

IS_COMP = "Cinnamic acid-d6"
# (시트명, 메서드, CSV 후보, 합산할 추가 화합물 후보 목록)
# 합산 후보가 있으면 해당 화합물들의 Area를 더해서 입력
AS_COMPOUNDS = [
    ("Loganin",         "B", ["Loganin"],         []),
    ("Ginsenoside Rb1", "A", ["Ginsenoside Rb1"],  []),
    ("Decursin",        "B", ["Decursin"],         ["Decursin Angelate", "Decursin angelate"]),
]


def _c(ws, row, col, value="", fill=None, bold=False, num_fmt=None):
    from openpyxl.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return cell
    cell.value = value
    cell.font = Font(name="맑은 고딕", size=FS, bold=bold)
    cell.alignment = AC
    cell.border = BORDER
    if fill:    cell.fill = fill
    if num_fmt: cell.number_format = num_fmt
    return cell


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _title(ws, row, text, c1, c2):
    _merge(ws, row, c1, row, c2)
    cell = ws.cell(row=row, column=c1)
    cell.value = text
    cell.font = Font(name="맑은 고딕", size=FS + 1, bold=True, color="FFFFFF")
    cell.fill = TITLE_F
    cell.alignment = AL
    cell.border = BORDER
    ws.row_dimensions[row].height = 20


def _lookup(parsed, candidates):
    for name in candidates:
        if name in parsed:
            return parsed[name]
    for cname in parsed:
        for cand in candidates:
            if cand.lower() in cname.lower():
                return parsed[cname]
    return {}


def _get_is(parsed):
    for name in parsed:
        if "cinnamic" in name.lower():
            return parsed[name]
    return {}


def _stats_rows(ws, row, val_start, val_end, cols):
    for label, func in [("평균", "AVERAGE"), ("표준편차", "STDEV"), ("%RSD", None)]:
        _c(ws, row, 1, label, fill=GRAY, bold=True)
        for col in cols:
            cl = get_column_letter(col)
            rng = f"{cl}{val_start}:{cl}{val_end}"
            if func:
                formula = f"=IFERROR({func}({rng}),\"\")"
                fmt = "0.000" if col in (2, 4) else "0"
            else:
                ar, sr = row - 2, row - 1
                formula = f"=IFERROR(IF({cl}{ar}=0,\"\",{cl}{sr}/{cl}{ar}*100),\"\")"
                fmt = "0.00"
            cell = ws.cell(row, col)
            cell.value = formula
            cell.fill = GRAY
            cell.border = BORDER
            cell.font = Font(name="맑은 고딕", size=FS)
            cell.alignment = AC
            cell.number_format = fmt
        row += 1
    return row


def _sum_areas(base_data, extra_candidates, parsed, section):
    """base_data의 run 목록에서 extra 화합물의 area를 더해 새 run 목록 반환."""
    runs = base_data.get(section, [])
    if not extra_candidates:
        return runs
    extra = _lookup(parsed, extra_candidates)
    if not extra:
        return runs
    extras = [extra]
    result = []
    for i, run in enumerate(runs):
        base_area = run.get("area") or 0
        extra_area = 0
        for ex in extras:
            ex_runs = ex.get(section, [])
            if i < len(ex_runs):
                extra_area += (ex_runs[i].get("area") or 0)
        new_run = dict(run)
        new_run["area"] = base_area + extra_area if (base_area or extra_area) else None
        result.append(new_run)
    return result


def _sum_sp_lots(base_data, extra_candidates, parsed, method):
    """SP lot 그룹화 시 extra 화합물 area를 합산."""
    def _filter(runs, m):
        needle = f"_{m}-".lower()
        return [r for r in runs if needle in r.get("name", "").lower()]

    def _group(runs):
        lots = {}
        for r in runs:
            m = re.search(r"^(.+?)_([ABab])-(\d+)$", r.get("name", ""))
            if m:
                lot, rn = m.group(1), int(m.group(3))
                lots.setdefault(lot, {})[rn] = r
        return lots

    base_runs = _filter(base_data.get("sp", []), method)
    base_lots = _group(base_runs)

    extra = _lookup(parsed, extra_candidates)
    extras = [extra] if extra else []
    extra_lot_maps = []
    for ex in extras:
        ex_runs = _filter(ex.get("sp", []), method)
        extra_lot_maps.append(_group(ex_runs))

    all_lots = sorted(set(list(base_lots.keys())))
    result = {}
    for lot in all_lots:
        base_lot_data = base_lots.get(lot, {})
        all_rns = set(base_lot_data.keys())
        for em in extra_lot_maps:
            all_rns |= set(em.get(lot, {}).keys())
        for rn in sorted(all_rns):
            base_area = (base_lot_data.get(rn) or {}).get("area") or 0
            extra_area = 0
            for em in extra_lot_maps:
                extra_area += ((em.get(lot, {}).get(rn) or {}).get("area") or 0)
            result.setdefault(lot, {})[rn] = base_area + extra_area if (base_area or extra_area) else None
    return result


def _build_sheet(ws, comp_name, method, csv_candidates, extra_candidates, parsed):
    for col, w in {1: 16, 2: 10, 3: 14, 4: 10, 5: 14, 6: 10}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    comp_data = _lookup(parsed, csv_candidates)
    is_data   = _get_is(parsed)
    # STD: extra 화합물 area 합산 (Decursin + Decursin Angelate 등)
    std_runs  = _sum_areas(comp_data, extra_candidates, parsed, "std")
    is_std    = is_data.get("std", [])

    row = 1
    _title(ws, row, f"표준액 정보 - {comp_name}", 1, 6); row += 1

    _c(ws, row, 1, "구분",   fill=BLUE_H, bold=True)
    _c(ws, row, 2, comp_name, fill=BLUE_H, bold=True); _merge(ws, row, 2, row, 3)
    _c(ws, row, 4, IS_COMP,   fill=BLUE_H, bold=True); _merge(ws, row, 4, row, 5)
    _c(ws, row, 6, "STD/IS",  fill=BLUE_H, bold=True); row += 1

    _c(ws, row, 1, "", fill=BLUE_H)
    for col, lbl in [(2, "RT"), (3, "Area"), (4, "RT"), (5, "Area"), (6, "비율")]:
        _c(ws, row, col, lbl, fill=BLUE_H, bold=True)
    row += 1

    val_start = row
    for i in range(6):
        _c(ws, row, 1, "Value" if i == 0 else "", fill=WHITE, bold=(i == 0))
        std = std_runs[i] if i < len(std_runs) else {}
        ist = is_std[i]   if i < len(is_std)   else {}
        _c(ws, row, 2, std.get("rt"),   fill=YELLOW, num_fmt="0.000")
        _c(ws, row, 3, std.get("area"), fill=YELLOW, num_fmt="0")
        _c(ws, row, 4, ist.get("rt"),   fill=YELLOW, num_fmt="0.000")
        _c(ws, row, 5, ist.get("area"), fill=YELLOW, num_fmt="0")
        c3, c5 = get_column_letter(3), get_column_letter(5)
        cell = ws.cell(row, 6)
        cell.value = f"=IFERROR({c3}{row}/{c5}{row},\"\")"
        cell.fill = GRAY; cell.border = BORDER
        cell.font = Font(name="맑은 고딕", size=FS)
        cell.alignment = AC; cell.number_format = "0.0000"
        row += 1
    val_end = row - 1
    if val_start < val_end:
        _merge(ws, val_start, 1, val_end, 1)

    row = _stats_rows(ws, row, val_start, val_end, [2, 3, 4, 5, 6])
    row += 1

    is_sp_all = is_data.get("sp", [])

    def _filter(runs, m):
        needle = f"_{m}-".lower()
        return [r for r in runs if needle in r.get("name", "").lower()]

    def _group(runs):
        lots = {}
        for r in runs:
            m = re.search(r"^(.+?)_([ABab])-(\d+)$", r.get("name", ""))
            if m:
                lot, rn = m.group(1), int(m.group(3))
                lots.setdefault(lot, {})[rn] = r.get("area")
        return lots

    # SP: extra 화합물 area 합산 후 lot 그룹화
    sp_lots  = _sum_sp_lots(comp_data, extra_candidates, parsed, method)
    is_runs  = _filter(is_sp_all, method)
    is_lots  = _group(is_runs)
    all_lots = sorted(set(list(sp_lots.keys()) + list(is_lots.keys()))) or ["Sample 1"]
    n  = len(all_lots)
    dc = 2

    def _short(lot):
        return re.sub(r"^.+?_LT_", "LT_", lot) if "_LT_" in lot else lot

    for lots_data, title in [(sp_lots, f"검액 정보 - {comp_name}"),
                              (is_lots, f"IS 검액 정보 - {IS_COMP}")]:
        for col in range(dc, dc + n):
            ws.column_dimensions[get_column_letter(col)].width = 13
        _title(ws, row, title, 1, dc + n - 1); row += 1

        _c(ws, row, 1, "구분", fill=BLUE_H, bold=True)
        for si, lot in enumerate(all_lots):
            _c(ws, row, dc + si, _short(lot), fill=BLUE_H, bold=True)
        row += 1

        _c(ws, row, 1, "", fill=BLUE_H)
        for si in range(n):
            _c(ws, row, dc + si, "Area", fill=BLUE_H, bold=True)
        row += 1

        v_start = row
        for ri in range(3):
            _c(ws, row, 1, "Value" if ri == 0 else "", fill=WHITE, bold=(ri == 0))
            for si, lot in enumerate(all_lots):
                area = lots_data.get(lot, {}).get(ri + 1) if lots_data else None
                _c(ws, row, dc + si, area, fill=YELLOW, num_fmt="0")
            row += 1
        v_end = row - 1
        if v_start < v_end:
            _merge(ws, v_start, 1, v_end, 1)
        row = _stats_rows(ws, row, v_start, v_end, list(range(dc, dc + n)))
        row += 1


def write_gj_as_result(parsed: dict) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for comp_name, method, csv_candidates, extra_candidates in AS_COMPOUNDS:
        ws = wb.create_sheet(comp_name[:31])
        _build_sheet(ws, comp_name, method, csv_candidates, extra_candidates, parsed)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
