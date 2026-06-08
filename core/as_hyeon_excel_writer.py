# -*- coding: utf-8 -*-
"""
AS 함량 결과 엑셀 생성기 — 현탁제용
현탁제 화합물: Amygdalin / Paeoniflorin / Cinnamic acid / Glycyrrhzin
STD/I.S 열 포함 / 검액 A·B 색상 구분
"""

import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 색상/스타일 상수 ──────────────────────────────────────────────
YELLOW   = PatternFill("solid", fgColor="FFFFE1")
BLUE_H   = PatternFill("solid", fgColor="BDD7EE")
GRAY     = PatternFill("solid", fgColor="D9D9D9")
WHITE    = PatternFill("solid", fgColor="FFFFFF")
A_HEADER = PatternFill("solid", fgColor="9DC3E6")   # 검액 A 헤더
B_HEADER = PatternFill("solid", fgColor="A9D18E")   # 검액 B 헤더
A_DATA   = PatternFill("solid", fgColor="DAEEF3")   # 검액 A 데이터
B_DATA   = PatternFill("solid", fgColor="E2EFDA")   # 검액 B 데이터

FONT_NORMAL = Font(name="맑은 고딕", size=14)
FONT_BOLD   = Font(name="맑은 고딕", size=14, bold=True)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left",   vertical="center")
THIN    = Side(style="thin")
BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _c(ws, row, col, value="", fill=None, bold=False, align="center", num_fmt=None):
    from openpyxl.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return cell
    cell.value      = value
    cell.font       = FONT_BOLD if bold else FONT_NORMAL
    cell.alignment  = ALIGN_C if align == "center" else ALIGN_L
    cell.border     = BORDER
    if fill:    cell.fill = fill
    if num_fmt: cell.number_format = num_fmt
    return cell


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _row_height(ws, row, h):
    ws.row_dimensions[row].height = h


# ── 메인 ─────────────────────────────────────────────────────────
def write_hyeon_result(parsed: dict, lot_groups: dict) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    IS_COMP   = "Cinnamic acid-d6"
    COMP_GRP1 = ["Amygdalin", "Paeoniflorin", "Cinnamic acid"]
    COMP_GRP2 = ["Glycyrrhzin"]
    SP_COMPS  = ["Amygdalin", "Paeoniflorin", "Cinnamic acid", "Glycyrrhzin"]

    std_runs     = parsed.get(IS_COMP, {}).get("std", [])
    is_std_areas = [r["area"] for r in std_runs]

    for lot_name, lot_data in lot_groups.items():
        ws = wb.create_sheet(title=lot_name[:31])
        _build_sheet(ws, parsed, is_std_areas, lot_data,
                     IS_COMP, COMP_GRP1, COMP_GRP2, SP_COMPS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_sheet(ws, parsed, is_std_areas, lot_data,
                 IS_COMP, COMP_GRP1, COMP_GRP2, SP_COMPS):

    # 열 너비: 구분(1) + 화합물당 2열(RT/Response) × 최대 3화합물 = 7열
    _col_widths(ws, {
        1:  12,   # 구분
        2:  10,   # RT
        3:  14,   # Response
        4:  10,
        5:  14,
        6:  10,
        7:  14,
        8:  20,   # SP 화합물 열
        9:  20,
        10: 20,
        11: 20,
    })

    row = 1

    # ══════════════════════════════════════════════
    # 4) 표준액 정보
    # ══════════════════════════════════════════════
    _c(ws, row, 1, "4) 표준액 정보", bold=True)
    _merge(ws, row, 1, row, 11)
    _row_height(ws, row, 22)
    row += 1

    row = _is_table(ws, row, parsed.get(IS_COMP, {}), IS_COMP)
    row += 1

    row = _compound_group_table(ws, row, COMP_GRP1, parsed, is_std_areas)
    row += 1

    row = _compound_group_table(ws, row, COMP_GRP2, parsed, is_std_areas)
    row += 1

    # ══════════════════════════════════════════════
    # 5) 검액 정보
    # ══════════════════════════════════════════════
    _c(ws, row, 1, "5) 검액 정보", bold=True)
    _merge(ws, row, 1, row, 11)
    _row_height(ws, row, 22)
    row += 1

    row = _sp_is_table(ws, row, IS_COMP, lot_data)
    row += 1
    row = _sp_compound_table(ws, row, SP_COMPS, lot_data)
    row += 1


# ── IS 표 ─────────────────────────────────────────────────────────
def _is_table(ws, row, is_data, comp_name):
    std_rows = is_data.get("std", [])

    _c(ws, row, 1, "구분",    fill=BLUE_H, bold=True)
    _c(ws, row, 2, comp_name, fill=BLUE_H, bold=True)
    _merge(ws, row, 2, row, 3)
    row += 1

    _c(ws, row, 1, "",          fill=BLUE_H)
    _c(ws, row, 2, "RT",        fill=BLUE_H, bold=True)
    _c(ws, row, 3, "Response",  fill=BLUE_H, bold=True)
    row += 1

    val_start = row
    for i in range(6):
        r = std_rows[i] if i < len(std_rows) else {}
        _c(ws, row, 1, "Value" if i == 0 else "", fill=WHITE, bold=(i == 0))
        _c(ws, row, 2, r.get("rt"),   fill=YELLOW, num_fmt="0.000")
        _c(ws, row, 3, r.get("area"), fill=YELLOW, num_fmt="0")
        row += 1
    if val_start < row - 1:
        _merge(ws, val_start, 1, row - 1, 1)

    _c(ws, row, 1, "평균", fill=GRAY, bold=True)
    for col, fmt in [(2, "0.000"), (3, "0")]:
        cl = get_column_letter(col)
        cell = ws.cell(row, col, f'=IFERROR(AVERAGE({cl}{val_start}:{cl}{row-1}),"")' )
        cell.fill = GRAY; cell.border = BORDER
        cell.font = FONT_NORMAL; cell.alignment = ALIGN_C
        cell.number_format = fmt
    row += 1

    return row


# ── 화합물 그룹 표 (RT + Response만) ─────────────────────────────
def _compound_group_table(ws, row, compounds, parsed, is_std_areas):
    """화합물당 2열: RT / Response"""
    COL_STARTS = [2 + i * 2 for i in range(len(compounds))]

    # 화합물 헤더
    _c(ws, row, 1, "구분", fill=BLUE_H, bold=True)
    for i, comp in enumerate(compounds):
        cs = COL_STARTS[i]
        _c(ws, row, cs,   comp, fill=BLUE_H, bold=True)
        _c(ws, row, cs+1, "",   fill=BLUE_H)
        _merge(ws, row, cs, row, cs + 1)
    row += 1

    # 서브헤더
    _c(ws, row, 1, "", fill=BLUE_H)
    for cs in COL_STARTS:
        _c(ws, row, cs,   "RT",       fill=BLUE_H, bold=True)
        _c(ws, row, cs+1, "Response", fill=BLUE_H, bold=True)
    row += 1

    # Value 6행
    val_start = row
    for i in range(6):
        _c(ws, row, 1, "Value" if i == 0 else "", fill=WHITE, bold=(i == 0))
        for j, comp in enumerate(compounds):
            cs = COL_STARTS[j]
            comp_data = parsed.get(comp, {}).get("std", [])
            r = comp_data[i] if i < len(comp_data) else {}
            _c(ws, row, cs,   r.get("rt"),   fill=YELLOW, num_fmt="0.000")
            _c(ws, row, cs+1, r.get("area"), fill=YELLOW, num_fmt="0")
        row += 1

    if val_start < row - 1:
        _merge(ws, val_start, 1, row - 1, 1)

    # 통계 3행
    for stat_label, stat_func in [("평균", "AVERAGE"), ("표준편차", "STDEV"), ("%RSD", None)]:
        _c(ws, row, 1, stat_label, fill=GRAY, bold=True)
        for cs in COL_STARTS:
            resp_col = get_column_letter(cs + 1)
            rng = f"{resp_col}{val_start}:{resp_col}{val_start+5}"
            if stat_func:
                formula = f'=IFERROR({stat_func}({rng}),"")'
            else:
                avg_r = row - 2; std_r = row - 1
                formula = f'=IFERROR(IF({resp_col}{avg_r}=0,"",{resp_col}{std_r}/{resp_col}{avg_r}*100),"")'
            cell = ws.cell(row, cs + 1, formula)
            cell.fill = GRAY; cell.border = BORDER
            cell.font = FONT_NORMAL; cell.alignment = ALIGN_C
            cell.number_format = "0.00"
            if stat_label == "평균":
                rt_col = get_column_letter(cs)
                cell = ws.cell(row, cs, f'=IFERROR(AVERAGE({rt_col}{val_start}:{rt_col}{val_start+5}),"")' )
                cell.fill = GRAY; cell.border = BORDER
                cell.font = FONT_NORMAL; cell.alignment = ALIGN_C
                cell.number_format = "0.000"
            else:
                _c(ws, row, cs, fill=GRAY)
        row += 1

    return row


# ── SP IS 표 ─────────────────────────────────────────────────────
def _sp_is_table(ws, row, is_comp, lot_data):
    a_runs = lot_data.get("A", [])
    b_runs = lot_data.get("B", [])

    _c(ws, row, 1, "구분\n(I.S)", fill=BLUE_H, bold=True)
    _c(ws, row, 2, f"{is_comp}\n검액 A", fill=A_HEADER, bold=True)
    _c(ws, row, 3, f"{is_comp}\n검액 B", fill=B_HEADER, bold=True)
    _row_height(ws, row, 36)
    row += 1

    _c(ws, row, 1, "",         fill=BLUE_H)
    _c(ws, row, 2, "Response", fill=A_HEADER, bold=True)
    _c(ws, row, 3, "Response", fill=B_HEADER, bold=True)
    row += 1

    for sp_idx, label in enumerate(["SP1", "SP2"]):
        a_area = a_runs[sp_idx].get(is_comp) if sp_idx < len(a_runs) else None
        b_area = b_runs[sp_idx].get(is_comp) if sp_idx < len(b_runs) else None
        _c(ws, row, 1, label, fill=BLUE_H, bold=True)
        _c(ws, row, 2, a_area, fill=YELLOW, num_fmt="0")
        _c(ws, row, 3, b_area, fill=YELLOW, num_fmt="0")
        row += 1

    return row


# ── SP 화합물 표 ─────────────────────────────────────────────────
def _sp_compound_table(ws, row, sp_comps, lot_data):
    a_runs = lot_data.get("A", [])
    b_runs = lot_data.get("B", [])

    # A run: Amygdalin + Cinnamic acid / B run: Paeoniflorin + Glycyrrhzin
    A_COMPS = {"Amygdalin", "Cinnamic acid"}
    B_COMPS = {"Paeoniflorin", "Glycyrrhzin"}

    def _hdr(comp):  return A_HEADER if comp in A_COMPS else B_HEADER
    def _data(comp): return A_DATA   if comp in A_COMPS else B_DATA

    # 헤더
    _c(ws, row, 1, "구분", fill=BLUE_H, bold=True)
    for ci, comp in enumerate(sp_comps, 2):
        _c(ws, row, ci, comp, fill=_hdr(comp), bold=True)
    _row_height(ws, row, 40)
    row += 1

    _c(ws, row, 1, "", fill=BLUE_H)
    for ci, comp in enumerate(sp_comps, 2):
        _c(ws, row, ci, "Response", fill=_hdr(comp), bold=True)
    row += 1

    for sp_idx, label in enumerate(["SP1", "SP2"]):
        _c(ws, row, 1, label, fill=BLUE_H, bold=True)
        for ci, comp in enumerate(sp_comps, 2):
            runs = a_runs if comp in A_COMPS else b_runs
            area = runs[sp_idx].get(comp) if sp_idx < len(runs) else None
            _c(ws, row, ci, area, fill=YELLOW, num_fmt="0")
        row += 1

    return row
