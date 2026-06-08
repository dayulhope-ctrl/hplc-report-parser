# -*- coding: utf-8 -*-
import sys, os, tempfile, shutil, io
from pathlib import Path
import streamlit as st

# 아미노산 분석 모듈 경로 추가
AA_DIR = Path(r"C:\Users\user\Downloads\파워라센 아미노산 함량")
sys.path.insert(0, str(AA_DIR))

from parse_pdf import parse_both, validate_data
from fill_form import fill_form
from template_creator import build_lot_sheet, build_standard_area_sheet
import openpyxl
import json

TEMPLATE = AA_DIR / "아미노산_함량_분석_양식.xlsx"

st.set_page_config(page_title="아미노산 함량 분석", layout="wide")
st.title("아미노산 함량 분석 (파워라센)")
st.markdown("---")

# ── 파일 업로드 ───────────────────────────────────────────────────────────────
st.subheader("PDF 파일 업로드")
col1, col2 = st.columns(2)
with col1:
    std_file = st.file_uploader("표준액 PDF (STD.pdf)", type="pdf", key="aa_std")
with col2:
    sp_file  = st.file_uploader("검액 PDF (SP.pdf)",   type="pdf", key="aa_sp")

if not TEMPLATE.exists():
    st.error(f"엑셀 양식 파일이 없습니다: {TEMPLATE}\n"
             f"먼저 template_creator.py를 실행해주세요.")
    st.stop()

if st.button("분석 실행", type="primary", disabled=not (std_file and sp_file)):

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)

        # PDF 임시 저장
        std_path = tmpdir / "STD.pdf"
        sp_path  = tmpdir / "SP.pdf"
        std_path.write_bytes(std_file.read())
        sp_path.write_bytes(sp_file.read())

        # ── STEP 1: PDF 파싱 ──────────────────────────────────────────────────
        with st.spinner("PDF 파싱 중..."):
            try:
                data = parse_both(str(std_path), str(sp_path))
            except Exception as e:
                st.error(f"PDF 파싱 오류: {e}")
                st.stop()

        warns = validate_data(data)
        if warns:
            with st.expander(f"경고 {len(warns)}건", expanded=True):
                for w in warns[:30]:
                    st.warning(w)
        else:
            st.success("데이터 검증 통과")

        lot_names = sorted(data["lots"].keys())
        st.info(f"감지된 Lot: {', '.join(lot_names)} ({len(lot_names)}개)")

        # ── STEP 2: 엑셀 생성 ─────────────────────────────────────────────────
        with st.spinner("엑셀 생성 중..."):
            out_path = tmpdir / "결과.xlsx"
            shutil.copy2(TEMPLATE, out_path)

            wb = openpyxl.load_workbook(str(out_path))

            # 표준액(면적) 시트에서 avg 셀 주소 추출
            from fill_form import find_std_area_cells as _find_std
            from openpyxl.utils import get_column_letter
            std_avg_cells = {}
            if "표준액(면적)" in wb.sheetnames:
                ws_std = wb["표준액(면적)"]
                avg_map = _find_std(ws_std)
                for aa, addrs in avg_map.items():
                    if addrs:
                        last = addrs[-1]
                        col  = ''.join(filter(str.isalpha, last))
                        row  = int(''.join(filter(str.isdigit, last))) + 1
                        std_avg_cells[aa] = f"{col}{row}"

            # lot 시트 생성
            for lot in lot_names:
                if lot not in wb.sheetnames:
                    build_lot_sheet(wb, lot, std_avg_cells)

            wb.save(str(out_path))

            # 데이터 입력
            json_tmp = tmpdir / "_data.json"
            with open(json_tmp, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            try:
                fill_form(str(out_path), str(json_tmp), str(out_path))
            except Exception as e:
                st.error(f"엑셀 입력 오류: {e}")
                st.stop()

        # ── 결과 다운로드 ──────────────────────────────────────────────────────
        st.success(f"완료! Lot {len(lot_names)}개 결과 생성됨")

        excel_bytes = out_path.read_bytes()

    st.download_button(
        label="결과 엑셀 다운로드",
        data=excel_bytes,
        file_name=f"아미노산_함량_결과.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
