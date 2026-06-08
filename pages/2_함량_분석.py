# -*- coding: utf-8 -*-
import sys, os, tempfile, math
from pathlib import Path
import streamlit as st
import pandas as pd
import io, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))
from core.as_parser import parse_as_summary

st.set_page_config(page_title="함량(AS) 분석", layout="wide")
st.title("함량 (AS) 분석")
st.markdown("---")

# ── 제품 선택 ──────────────────────────────────────────────────
product_name = st.selectbox("제품", ["우황청심원 환제", "우황청심원 현탁제"])

# ── PDF 업로드 ──────────────────────────────────────────────────
st.subheader("AS Summary PDF 업로드")
as_file = st.file_uploader("AS_Summary.pdf", type="pdf")

if st.button("분석 실행", type="primary", disabled=as_file is None):

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(as_file.read())
        tmp_path = tmp.name

    try:
        with st.spinner("PDF 파싱 중..."):
            parsed = parse_as_summary(tmp_path)
    except Exception as e:
        st.error(f"파싱 오류: {e}")
        st.stop()
    finally:
        os.unlink(tmp_path)

    if not parsed:
        st.error("PDF에서 데이터를 추출하지 못했습니다.")
        st.stop()

    st.success(f"추출 완료: {len(parsed)}개 항목")

    # ── 엑셀 생성 ──────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
    GRAY_FILL   = PatternFill("solid", fgColor="D9D9D9")
    BOLD        = Font(bold=True)
    CENTER      = Alignment(horizontal="center", vertical="center")
    BORDER      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def hdr(ws, row, col, val, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = BOLD
        c.alignment = CENTER
        c.border = BORDER
        if fill: c.fill = fill
        return c

    def dat(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = CENTER
        c.border = BORDER
        return c

    # 각 compound별 시트 생성
    for comp_name, comp_data in parsed.items():
        safe_name = comp_name[:31].replace("/","_").replace("\\","_").replace("*","_") \
                                  .replace("?","_").replace("[","_").replace("]","_").replace(":","_")
        ws = wb.create_sheet(title=safe_name)
        ws.column_dimensions["A"].width = 25
        for col in ["B","C","D","E"]:
            ws.column_dimensions[col].width = 15

        row = 1
        hdr(ws, row, 1, comp_name, fill=HEADER_FILL)
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

        for section_key, section_label in [
            ("system_check", "System Check"),
            ("stability",    "Stability"),
            ("std",          "STD"),
            ("sp",           "검액 (SP)"),
        ]:
            rows_data = comp_data.get(section_key, [])
            if not rows_data:
                continue

            hdr(ws, row, 1, section_label, fill=GRAY_FILL)
            ws.merge_cells(f"A{row}:E{row}")
            row += 1

            hdr(ws, row, 1, "Name",     fill=GRAY_FILL)
            hdr(ws, row, 2, "RT",       fill=GRAY_FILL)
            hdr(ws, row, 3, "Response", fill=GRAY_FILL)
            hdr(ws, row, 4, "S/N",      fill=GRAY_FILL)
            hdr(ws, row, 5, "File",     fill=GRAY_FILL)
            row += 1

            for entry in rows_data:
                dat(ws, row, 1, entry.get("name"))
                dat(ws, row, 2, entry.get("rt"))
                dat(ws, row, 3, entry.get("resp"))
                dat(ws, row, 4, entry.get("sn"))
                dat(ws, row, 5, entry.get("file"))
                row += 1

            # STD 구간은 평균/SD/%RSD 추가
            if section_key == "std" and rows_data:
                resps = [r["resp"] for r in rows_data if r.get("resp") is not None]
                if resps:
                    avg  = sum(resps) / len(resps)
                    sd   = math.sqrt(sum((x-avg)**2 for x in resps)/(len(resps)-1)) if len(resps)>1 else 0
                    rsd  = sd/avg*100 if avg else 0
                    for label, val in [("평균", avg), ("SD", sd), ("%RSD", rsd)]:
                        c = ws.cell(row=row, column=1, value=label)
                        c.font = BOLD; c.alignment = CENTER; c.border = BORDER; c.fill = GRAY_FILL
                        dat(ws, row, 3, round(val, 3))
                        row += 1

            row += 1  # 섹션 간 빈 행

    # 요약 시트 (전체 STD 평균 Response 한눈에)
    ws_sum = wb.create_sheet(title="요약", index=0)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 15
    ws_sum.column_dimensions["D"].width = 15

    hdr(ws_sum, 1, 1, "성분",         fill=HEADER_FILL)
    hdr(ws_sum, 1, 2, "STD 평균 RT",  fill=HEADER_FILL)
    hdr(ws_sum, 1, 3, "STD 평균 Resp",fill=HEADER_FILL)
    hdr(ws_sum, 1, 4, "SP 시료 수",   fill=HEADER_FILL)

    for r_idx, (comp_name, comp_data) in enumerate(parsed.items(), start=2):
        std_rows = comp_data.get("std", [])
        sp_rows  = comp_data.get("sp",  [])

        avg_rt   = round(sum(r["rt"]   for r in std_rows)/len(std_rows), 3) if std_rows else None
        avg_resp = round(sum(r["resp"] for r in std_rows)/len(std_rows), 1) if std_rows else None

        dat(ws_sum, r_idx, 1, comp_name)
        dat(ws_sum, r_idx, 2, avg_rt)
        dat(ws_sum, r_idx, 3, avg_resp)
        dat(ws_sum, r_idx, 4, len(sp_rows))

    # 엑셀 bytes 반환
    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    # ── 미리보기 ────────────────────────────────────────────────
    st.subheader("요약 미리보기")
    preview = []
    for comp_name, comp_data in parsed.items():
        std_rows = comp_data.get("std", [])
        sp_rows  = comp_data.get("sp",  [])
        avg_resp = round(sum(r["resp"] for r in std_rows)/len(std_rows),1) if std_rows else None
        preview.append({
            "성분": comp_name,
            "STD 런 수": len(std_rows),
            "STD 평균 Response": avg_resp,
            "SP 시료 수": len(sp_rows),
        })
    st.dataframe(pd.DataFrame(preview), use_container_width=True, hide_index=True)

    st.download_button(
        label="결과 엑셀 다운로드",
        data=excel_bytes,
        file_name=f"{product_name}_함량분석결과.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
